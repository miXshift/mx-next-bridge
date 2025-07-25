import openpyxl
import pandas as pd
from openpyxl.utils import get_column_letter
import json

def analyze_excel_formulas_detailed(file_path):
    """Analyze formulas in Excel file tabs with more detail"""
    # Load the workbook
    wb = openpyxl.load_workbook(file_path, data_only=False)
    
    results = {}
    
    # Analyze p1_campaign_kpis and p2_campaign_kpis tabs
    for sheet_name in ['p1_campaign_kpis', 'p2_campaign_kpis']:
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            print(f"\n{'='*80}")
            print(f"Analyzing sheet: {sheet_name}")
            print('='*80)
            
            sheet_info = {
                'formulas': {},
                'headers': {},
                'column_mappings': {},
                'aggregation_rules': {},
                'dimensions': f"Rows: {ws.max_row}, Columns: {ws.max_column}"
            }
            
            # Find the actual header row (might not be row 1)
            header_row = None
            for row in range(1, min(10, ws.max_row + 1)):
                cell_value = ws.cell(row=row, column=1).value
                if cell_value and isinstance(cell_value, str) and 'Campaign' in str(cell_value):
                    header_row = row
                    break
            
            if header_row:
                print(f"\nHeaders found at row {header_row}:")
                # Get headers
                for col in range(1, ws.max_column + 1):
                    cell = ws.cell(row=header_row, column=col)
                    if cell.value:
                        col_letter = get_column_letter(col)
                        sheet_info['headers'][col_letter] = str(cell.value)
                        print(f"  Column {col_letter}: {cell.value}")
            
            # Analyze all cells for formulas
            print(f"\nAnalyzing all formulas in sheet:")
            formula_by_column = {}
            
            for row in range(1, ws.max_row + 1):
                for col in range(1, ws.max_column + 1):
                    cell = ws.cell(row=row, column=col)
                    if cell.value and isinstance(cell.value, str) and cell.value.startswith('='):
                        col_letter = get_column_letter(col)
                        cell_ref = f"{col_letter}{row}"
                        
                        # Group formulas by column
                        if col_letter not in formula_by_column:
                            formula_by_column[col_letter] = []
                        
                        formula_by_column[col_letter].append({
                            'cell': cell_ref,
                            'row': row,
                            'formula': cell.value,
                            'header': sheet_info['headers'].get(col_letter, 'No header')
                        })
            
            # Analyze formula patterns by column
            print(f"\nFormula patterns by column:")
            for col_letter, formulas in sorted(formula_by_column.items()):
                header = sheet_info['headers'].get(col_letter, 'No header')
                print(f"\n  Column {col_letter} ({header}):")
                
                # Show first few unique formulas
                unique_formulas = {}
                for f in formulas:
                    formula_pattern = f['formula']
                    # Try to generalize the formula pattern
                    import re
                    generalized = re.sub(r'\d+', 'N', formula_pattern)
                    if generalized not in unique_formulas:
                        unique_formulas[generalized] = f
                
                for pattern, example in list(unique_formulas.items())[:5]:
                    print(f"    Pattern: {pattern}")
                    print(f"    Example: {example['formula']} (from cell {example['cell']})")
                
                # Store column mapping info
                sheet_info['column_mappings'][col_letter] = {
                    'header': header,
                    'formula_count': len(formulas),
                    'unique_patterns': len(unique_formulas),
                    'sample_formulas': [f['formula'] for f in formulas[:3]]
                }
            
            # Look for aggregation formulas (SUM, AVERAGE, etc.)
            print(f"\nAggregation formulas found:")
            for col_letter, formulas in formula_by_column.items():
                for f in formulas:
                    if any(agg in f['formula'].upper() for agg in ['SUM', 'AVERAGE', 'COUNT', 'MAX', 'MIN']):
                        print(f"  {f['cell']}: {f['formula']}")
                        sheet_info['aggregation_rules'][f['cell']] = {
                            'formula': f['formula'],
                            'column': col_letter,
                            'header': sheet_info['headers'].get(col_letter, 'No header')
                        }
            
            results[sheet_name] = sheet_info
            
            # Show data sample
            print(f"\nData sample (rows {header_row if header_row else 1} to {min(header_row + 10 if header_row else 10, ws.max_row)}):")
            try:
                df = pd.read_excel(file_path, sheet_name=sheet_name, 
                                 skiprows=header_row-1 if header_row and header_row > 1 else None,
                                 nrows=10)
                print(df.to_string())
            except:
                print("  Could not read data sample")
    
    # Save detailed results
    with open('../data/kpi_formula_analysis_detailed.json', 'w') as f:
        json.dump(results, f, indent=2)
    
    print(f"\n\nDetailed analysis saved to ../data/kpi_formula_analysis_detailed.json")
    
    # Also check Campaign sheet for reference
    if 'Campaign' in wb.sheetnames:
        print(f"\n\nChecking Campaign sheet structure (for VLOOKUP references):")
        ws = wb['Campaign']
        print(f"  Dimensions: {ws.max_row} rows x {ws.max_column} columns")
        print(f"  First row headers:")
        for col in range(1, min(10, ws.max_column + 1)):
            cell = ws.cell(row=1, column=col)
            if cell.value:
                print(f"    Column {get_column_letter(col)}: {cell.value}")
    
    return results

if __name__ == "__main__":
    file_path = "../data/HydrapakUS_SPABridge_MoM_February 2025vsJanuary 2025.xlsx"
    results = analyze_excel_formulas_detailed(file_path)
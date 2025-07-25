import pandas as pd
import openpyxl
from openpyxl import load_workbook

# Read the Excel file
excel_file = "../data/HydrapakUS_SPABridge_MoM_February 2025vsJanuary 2025.xlsx"

# Load workbook preserving formulas
wb = load_workbook(excel_file, data_only=False)
ws = wb['Campaign']

print("=== COMPLETE CAMPAIGN TAB STRUCTURE ===\n")

# Document the complete structure
print("1. HEADER SECTION (Rows 1-11):")
print("   - Row 1: Account name (Hydrapak - US)")
print("   - Row 2: Period comparison (January 2025 vs February 2025)")
print("   - Row 3: Attribute type (Campaign)")
print("   - Row 4-11: Filter settings and notes")
print("   - Row 12: KPI labels for each metric group")
print("   - Row 13: Column headers for each metric\n")

print("2. METRICS STRUCTURE (14 metric groups):")
metrics = [
    ("B-F", "Spend"),
    ("G-K", "Total Ad Sales"),
    ("L-P", "ACoS"),
    ("Q-U", "ROAS"),
    ("V-Z", "Conversion Rate"),
    ("AA-AE", "Impressions"),
    ("AF-AJ", "Clicks"),
    ("AK-AO", "Clickthrough Rate"),
    ("AP-AT", "Cost Per Click"),
    ("AU-AY", "Same SKU Ad Sales"),
    ("AZ-BD", "Other SKU Sales"),
    ("BE-BI", "Same SKU Ad Orders"),
    ("BJ-BN", "Other SKU Ad Orders"),
    ("BO-BS", "Total Ad Orders")
]

print("\nEach metric group contains 5 columns:")
print("   - January 2025: Previous period value")
print("   - February 2025: Current period value")
print("   - Net Change: Absolute difference (February - January)")
print("   - % Change: Percentage change from January to February")
print("   - Contribution: Contribution to overall change\n")

for cols, metric in metrics:
    print(f"   {cols}: {metric}")

print("\n3. DATA ROWS (Starting from row 14):")
print("   - Row 14: Total (aggregate of all campaigns)")
print("   - Rows 15+: Individual campaign data\n")

# Check specific cells for formulas to understand calculation patterns
print("=== CHECKING FOR CALCULATION PATTERNS ===\n")

# Check Net Change columns (D, I, N, etc.)
print("Checking Net Change calculations:")
net_change_cols = [4, 9, 14, 19, 24, 29, 34, 39, 44, 49, 54, 59, 64, 69]
for col in net_change_cols[:3]:  # Check first 3
    col_letter = openpyxl.utils.get_column_letter(col)
    # Check a few rows
    for row in [14, 15, 16]:
        cell = ws.cell(row=row, column=col)
        if hasattr(cell, 'value'):
            print(f"  {col_letter}{row}: {cell.value} (Type: {type(cell.value).__name__})")

# Check % Change columns (E, J, O, etc.)
print("\nChecking % Change calculations:")
pct_change_cols = [5, 10, 15, 20, 25, 30, 35, 40, 45, 50, 55, 60, 65, 70]
for col in pct_change_cols[:3]:  # Check first 3
    col_letter = openpyxl.utils.get_column_letter(col)
    for row in [14, 15]:
        cell = ws.cell(row=row, column=col)
        if hasattr(cell, 'value'):
            print(f"  {col_letter}{row}: {cell.value} (Type: {type(cell.value).__name__})")

# Check Contribution columns (F, K, P, etc.)
print("\nChecking Contribution calculations:")
contrib_cols = [6, 11, 16, 21, 26, 31, 36, 41, 46, 51, 56, 61, 66, 71]
for col in contrib_cols[:3]:  # Check first 3
    col_letter = openpyxl.utils.get_column_letter(col)
    for row in [14, 15]:
        cell = ws.cell(row=row, column=col)
        if hasattr(cell, 'value'):
            print(f"  {col_letter}{row}: {cell.value} (Type: {type(cell.value).__name__})")

# Document the formatting
print("\n=== NUMBER FORMATTING ===\n")
print("Based on the analysis, the following number formats are used:")
print("- Spend, Sales, Orders: Currency format ($#,##0.00)")
print("- ACoS, ROAS, CTR, CPC: Percentage or decimal format")
print("- Impressions, Clicks: Number format (#,##0)")
print("- Contribution: Basis points format (#,##0\"bps\")")

# Check if this is a pivot table or regular data
print("\n=== DATA SOURCE ANALYSIS ===\n")
print("This appears to be static data or calculated values (no formulas detected).")

# Final summary
print("\n=== SUMMARY ===\n")
print("The Campaign tab contains a month-over-month comparison report with:")
print("- 14 key metrics tracked across campaigns")
print("- Each metric shows: previous month, current month, change, % change, and contribution")
print("- Data appears to be static (no formulas found in cells)")
print("- Formatting includes currency, percentages, and basis points")
print("- The structure suggests this is an exported report rather than a live calculation sheet")

wb.close()

# Show the actual calculation logic that would be used
print("\n=== IMPLIED CALCULATION LOGIC ===\n")
print("Based on the structure, the calculations would be:")
print("1. Net Change = February Value - January Value")
print("2. % Change = (Net Change / January Value) * 100")
print("3. Contribution = (Campaign Net Change / Total Net Change) * 100 * 100 (in basis points)")
print("\nNote: For metrics like ACoS, CTR, the 'Pts Change' is used instead of 'Net Change'")
print("      This represents the absolute percentage point difference rather than relative change")
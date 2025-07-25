import pandas as pd
import openpyxl
from openpyxl import load_workbook
import re

# Read the Excel file
excel_file = "../data/HydrapakUS_SPABridge_MoM_February 2025vsJanuary 2025.xlsx"

# Load workbook to check for formulas
wb = load_workbook(excel_file, data_only=False)
ws = wb['Campaign']

print("=== CAMPAIGN TAB DETAILED ANALYSIS ===\n")
print(f"Sheet dimensions: {ws.max_row} rows x {ws.max_column} columns\n")

# Find the actual header row by looking for specific patterns
print("=== SEARCHING FOR ACTUAL HEADERS ===\n")
header_row = None
for row in range(1, 20):  # Check first 20 rows
    row_values = []
    for col in range(1, min(ws.max_column + 1, 20)):
        cell_value = ws.cell(row=row, column=col).value
        if cell_value:
            row_values.append(str(cell_value))
    
    # Look for rows that might contain column headers
    if any('Campaign' in str(v) or 'Impressions' in str(v) or 'Clicks' in str(v) or 'Cost' in str(v) for v in row_values):
        print(f"Row {row}: {row_values[:10]}...")  # Show first 10 values
        if len([v for v in row_values if v and not v.startswith('Unnamed')]) > 5:
            header_row = row

print(f"\nLikely header row: {header_row}")

# Get all column headers with their positions
print("\n=== ALL COLUMN HEADERS ===\n")
if header_row:
    headers = {}
    for col in range(1, ws.max_column + 1):
        cell = ws.cell(row=header_row, column=col)
        if cell.value:
            col_letter = openpyxl.utils.get_column_letter(col)
            headers[col] = (col_letter, cell.value)
            print(f"Column {col} ({col_letter}): {cell.value}")

# Search for ALL formulas in the sheet
print("\n=== COMPREHENSIVE FORMULA SEARCH ===\n")
formula_map = {}
for row in range(1, ws.max_row + 1):
    for col in range(1, ws.max_column + 1):
        cell = ws.cell(row=row, column=col)
        if isinstance(cell.value, str) and cell.value.startswith('='):
            col_letter = openpyxl.utils.get_column_letter(col)
            cell_ref = f"{col_letter}{row}"
            formula_map[cell_ref] = cell.value

if formula_map:
    # Group formulas by column to identify patterns
    formula_by_column = {}
    for cell_ref, formula in formula_map.items():
        col = re.match(r'([A-Z]+)\d+', cell_ref).group(1)
        if col not in formula_by_column:
            formula_by_column[col] = []
        formula_by_column[col].append((cell_ref, formula))
    
    # Show formula patterns by column
    for col in sorted(formula_by_column.keys()):
        formulas = formula_by_column[col]
        print(f"\nColumn {col}:")
        # Show first few formulas to identify pattern
        for i, (cell_ref, formula) in enumerate(formulas[:3]):
            print(f"  {cell_ref}: {formula}")
        if len(formulas) > 3:
            print(f"  ... and {len(formulas) - 3} more formulas in this column")
else:
    print("No formulas found in the sheet")

# Check cell formatting
print("\n=== CELL FORMATTING ANALYSIS ===\n")
format_samples = {}
for row in range(header_row or 1, min(ws.max_row + 1, header_row + 10) if header_row else 10):
    for col in range(1, min(ws.max_column + 1, 15)):
        cell = ws.cell(row=row, column=col)
        if cell.number_format and cell.number_format != 'General':
            col_letter = openpyxl.utils.get_column_letter(col)
            format_type = cell.number_format
            if format_type not in format_samples:
                format_samples[format_type] = []
            format_samples[format_type].append(f"{col_letter}{row}")

for format_type, cells in format_samples.items():
    print(f"Format: {format_type}")
    print(f"  Examples: {', '.join(cells[:5])}")
    if len(cells) > 5:
        print(f"  ... and {len(cells) - 5} more cells")

# Read with pandas to show actual data structure
print("\n=== DATA STRUCTURE ANALYSIS ===\n")
df = pd.read_excel(excel_file, sheet_name='Campaign', header=None)

# Find the actual data start row
data_start_row = None
for idx, row in df.iterrows():
    if idx > 20:  # Don't search too far
        break
    # Look for rows with actual campaign names or numeric data
    non_null_count = row.notna().sum()
    if non_null_count > 10:  # If more than 10 non-null values
        # Check if there are numeric values
        numeric_count = sum(1 for val in row if pd.api.types.is_numeric_dtype(type(val)) and pd.notna(val))
        if numeric_count > 5:
            data_start_row = idx
            break

print(f"Data appears to start at row: {data_start_row}")

# Show a sample of the actual data
if data_start_row and data_start_row > 0:
    df_with_headers = pd.read_excel(excel_file, sheet_name='Campaign', header=data_start_row-1)
    print(f"\nDataFrame with proper headers - Shape: {df_with_headers.shape}")
    print("\nColumn names:")
    for i, col in enumerate(df_with_headers.columns):
        print(f"  {i+1}. {col}")
    
    print("\nFirst few rows of actual data:")
    print(df_with_headers.head(3))

wb.close()
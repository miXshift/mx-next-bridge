#!/usr/bin/env python3
"""
Find the actual column headers for the Excel campaign data
"""

from openpyxl import load_workbook

def find_excel_headers():
    excel_path = "../data/HydrapakUS_SPABridge_MoM_February 2025vsJanuary 2025.xlsx"
    
    wb = load_workbook(excel_path, data_only=True)
    ws = wb['Campaign']
    
    print("SEARCHING FOR EXCEL HEADERS")
    print("="*60)
    
    # Look through first several rows to find headers
    for check_row in range(1, 15):
        print(f"\nRow {check_row} analysis:")
        row_values = []
        for col in range(1, min(21, ws.max_column + 1)):
            value = ws.cell(row=check_row, column=col).value
            row_values.append(value)
        
        # Check if this looks like a header row
        non_none_count = sum(1 for v in row_values if v is not None and str(v).strip() != '')
        print(f"  Non-empty values: {non_none_count}")
        
        if non_none_count > 10:  # Likely a header row
            print(f"  POTENTIAL HEADER ROW:")
            for i, value in enumerate(row_values, 1):
                if value is not None and str(value).strip() != '':
                    print(f"    Col {i}: '{value}'")
    
    # Look specifically at rows 12-14 which might contain the actual headers
    print(f"\n" + "="*60)
    print("DETAILED HEADER ANALYSIS (Rows 12-14):")
    
    for header_row in [12, 13]:
        print(f"\nRow {header_row} - Full content:")
        for col in range(1, ws.max_column + 1):
            value = ws.cell(row=header_row, column=col).value
            if value is not None and str(value).strip() != '':
                print(f"  Col {col}: '{value}'")
    
    # Now let's see the Skyflask data with potential headers
    skyflask_row = 33
    print(f"\n" + "="*60)
    print(f"SKYFLASK DATA (Row {skyflask_row}) with Row 13 as headers:")
    
    for col in range(1, min(16, ws.max_column + 1)):
        header = ws.cell(row=13, column=col).value
        value = ws.cell(row=skyflask_row, column=col).value
        print(f"  Col {col}: Header='{header}' | Value={value}")

if __name__ == "__main__":
    find_excel_headers()
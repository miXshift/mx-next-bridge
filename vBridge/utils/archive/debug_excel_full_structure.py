#!/usr/bin/env python3
"""
Debug script to fully understand Excel structure
"""

from openpyxl import load_workbook

def debug_full_excel_structure():
    excel_path = "../data/HydrapakUS_SPABridge_MoM_February 2025vsJanuary 2025.xlsx"
    
    wb = load_workbook(excel_path, data_only=True)
    ws = wb['Campaign']
    
    print("FULL EXCEL CAMPAIGN TAB ANALYSIS")
    print("="*60)
    print(f"Dimensions: {ws.max_row} rows x {ws.max_column} columns")
    
    # Show first 5 rows to understand header structure
    print(f"\nFirst 5 rows (first 15 columns):")
    for row in range(1, min(6, ws.max_row + 1)):
        print(f"\nRow {row}:")
        for col in range(1, min(16, ws.max_column + 1)):
            cell_value = ws.cell(row=row, column=col).value
            print(f"  Col {col}: '{cell_value}'")
    
    # Look for campaigns
    print(f"\nLooking for campaign rows (rows with campaign names):")
    campaign_rows = []
    for row in range(1, ws.max_row + 1):
        cell_value = ws.cell(row=row, column=1).value
        if cell_value and any(keyword in str(cell_value) for keyword in ['1000-', '10000-', 'Total']):
            campaign_rows.append((row, cell_value))
            if len(campaign_rows) <= 5:  # Show first 5
                print(f"  Row {row}: '{cell_value}'")
    
    print(f"\nFound {len(campaign_rows)} campaign rows")
    
    # Find Skyflask specifically
    skyflask_campaigns = []
    for row, campaign in campaign_rows:
        if 'Skyflask' in str(campaign):
            skyflask_campaigns.append((row, campaign))
    
    print(f"\nSkyflask campaigns found:")
    for row, campaign in skyflask_campaigns:
        print(f"  Row {row}: '{campaign}'")
        
        # Show data for this row
        print(f"    Data (first 10 values):")
        for col in range(1, min(11, ws.max_column + 1)):
            value = ws.cell(row=row, column=col).value
            print(f"      Col {col}: {value}")

if __name__ == "__main__":
    debug_full_excel_structure()
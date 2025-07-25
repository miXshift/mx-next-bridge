#!/usr/bin/env python3
"""
Detailed comparison with 10 decimal places to identify remaining discrepancies
"""

import pandas as pd
from openpyxl import load_workbook

def load_excel_campaign_row(excel_path, campaign_name):
    """Load specific campaign row from Excel"""
    wb = load_workbook(excel_path, data_only=True)
    ws = wb['Campaign']
    
    # Get headers
    headers = []
    for col in range(1, ws.max_column + 1):
        cell_value = ws.cell(row=1, column=col).value
        headers.append(cell_value if cell_value else f"Col_{col}")
    
    # Find campaign row
    for row in range(2, ws.max_row + 1):
        campaign_cell = ws.cell(row=row, column=1).value
        if campaign_cell and str(campaign_cell).strip() == campaign_name:
            campaign_data = []
            for col in range(1, ws.max_column + 1):
                cell_value = ws.cell(row=row, column=col).value
                campaign_data.append(cell_value)
            return dict(zip(headers, campaign_data))
    
    return None

def load_csv_campaign_row(csv_path, campaign_name):
    """Load specific campaign row from CSV with 10 decimal precision"""
    with open(csv_path, 'r') as f:
        lines = f.readlines()
    
    # Parse headers
    top_headers = lines[0].strip().split(',')
    sub_headers = lines[1].strip().split(',')
    
    # Create combined headers
    combined_headers = []
    for i, (top, sub) in enumerate(zip(top_headers, sub_headers)):
        if i == 0:
            combined_headers.append('Campaign Name')
        else:
            combined_headers.append(f"{top} {sub}")
    
    # Find campaign row
    for line in lines[2:]:
        if line.strip():
            row = line.strip().split(',')
            if row[0].strip() == campaign_name:
                campaign_data = {}
                for i, (header, value) in enumerate(zip(combined_headers, row)):
                    if i == 0:
                        campaign_data[header] = value
                    else:
                        try:
                            campaign_data[header] = float(value) if value != '' else 0.0
                        except ValueError:
                            campaign_data[header] = value
                return campaign_data
    
    return None

def analyze_with_precision():
    """Compare with high precision to identify remaining differences"""
    campaign_name = "1000-CONQ-NONBRAND-ASIN-Skyflask"
    excel_path = "../data/HydrapakUS_SPABridge_MoM_February 2025vsJanuary 2025.xlsx"
    csv_path = "../output/period_comparison.csv"
    
    print("HIGH PRECISION COMPARISON (10 decimal places)")
    print("=" * 80)
    print(f"Campaign: {campaign_name}")
    print("=" * 80)
    
    # Load data
    excel_data = load_excel_campaign_row(excel_path, campaign_name)
    csv_data = load_csv_campaign_row(csv_path, campaign_name)
    
    if not excel_data or not csv_data:
        print("ERROR: Campaign not found in both datasets")
        return
    
    # Define metrics to compare with focus on contribution values
    contribution_metrics = [
        ("Spend", "Contribution"),
        ("Total Ad Sales", "Contribution"),
        ("ACoS", "Contribution"),
        ("ROAS", "Contribution"),
        ("Conversion Rate", "Contribution"),
        ("Impressions", "Contribution"),
        ("Clicks", "Contribution"),
        ("CTR", "Contribution"),
        ("CPC", "Contribution"),
        ("Same SKU Ad Sales", "Contribution"),
        ("Other SKU Sales", "Contribution"),
        ("Same SKU Ad Orders", "Contribution"),
        ("Other SKU Ad Orders", "Contribution"),
        ("Total Ad Orders", "Contribution")
    ]
    
    print("\nCONTRIBUTION VALUES WITH HIGH PRECISION:")
    print("-" * 80)
    
    differences = []
    
    for metric, dimension in contribution_metrics:
        # Find matching columns
        excel_col = None
        for col in excel_data.keys():
            if metric in str(col) and dimension in str(col):
                excel_col = col
                break
        
        csv_col = f"{metric} {dimension}"
        
        if excel_col and csv_col in csv_data:
            excel_val = float(excel_data[excel_col]) if excel_data[excel_col] is not None else 0.0
            csv_val = float(csv_data[csv_col]) if csv_data[csv_col] is not None else 0.0
            
            diff = abs(excel_val - csv_val)
            
            # Show all values with high precision
            print(f"{metric:25} Excel: {excel_val:20.10f}")
            print(f"{' ':25} CSV:   {csv_val:20.10f}")
            print(f"{' ':25} Diff:  {diff:20.10f}")
            
            if diff > 0.0001:  # Significant difference
                print(f"{' ':25} *** SIGNIFICANT DIFFERENCE ***")
                differences.append((metric, excel_val, csv_val, diff))
            
            print()
    
    # Also check percentage changes for precision
    print("\nPERCENTAGE CHANGES WITH HIGH PRECISION:")
    print("-" * 80)
    
    pct_metrics = [
        ("Spend", "% Change"),
        ("Total Ad Sales", "% Change"),
        ("ACoS", "% Change"),
        ("ROAS", "% Change"),
        ("Conversion Rate", "% Change"),
        ("Clicks", "% Change"),
        ("CTR", "% Change"),
        ("CPC", "% Change")
    ]
    
    for metric, dimension in pct_metrics:
        excel_col = None
        for col in excel_data.keys():
            if metric in str(col) and dimension in str(col):
                excel_col = col
                break
        
        csv_col = f"{metric} {dimension}"
        
        if excel_col and csv_col in csv_data:
            excel_val = float(excel_data[excel_col]) if excel_data[excel_col] is not None else 0.0
            csv_val = float(csv_data[csv_col]) if csv_data[csv_col] is not None else 0.0
            
            diff = abs(excel_val - csv_val)
            
            if diff > 0.01:  # Only show if there's a meaningful difference
                print(f"{metric:25} Excel: {excel_val:20.10f}")
                print(f"{' ':25} CSV:   {csv_val:20.10f}")
                print(f"{' ':25} Diff:  {diff:20.10f}")
                print()
    
    # Summary of significant differences
    if differences:
        print("\nSUMMARY OF SIGNIFICANT CONTRIBUTION DIFFERENCES:")
        print("-" * 80)
        for metric, excel_val, csv_val, diff in sorted(differences, key=lambda x: x[3], reverse=True):
            print(f"{metric:25} Diff: {diff:15.10f} (Excel: {excel_val:10.2f}, CSV: {csv_val:10.2f})")
    
    # Debug: Show raw data for Spend Contribution
    print("\nDEBUG - Raw data for key metrics:")
    print("-" * 80)
    print("Spend Contribution:")
    excel_spend_contrib = None
    csv_spend_contrib = None
    for col in excel_data.keys():
        if "Spend" in str(col) and "Contribution" in str(col):
            excel_spend_contrib = excel_data[col]
            print(f"  Excel column '{col}': {excel_spend_contrib}")
            break
    
    if "Spend Contribution" in csv_data:
        csv_spend_contrib = csv_data["Spend Contribution"]
        print(f"  CSV value: {csv_spend_contrib}")

if __name__ == "__main__":
    analyze_with_precision()
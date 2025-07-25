#!/usr/bin/env python3
"""
Investigate why total row values differ significantly between Excel and CSV
"""

import pandas as pd
from openpyxl import load_workbook

def analyze_csv_source_data():
    """Analyze the source CSV data to understand what's being included"""
    csv_path = "../data/Hydrapak YTD - campaign.csv"
    
    print("CSV SOURCE DATA ANALYSIS")
    print("=" * 80)
    
    df = pd.read_csv(csv_path)
    df['Date'] = pd.to_datetime(df['DateKey'], format='%Y%m%d')
    
    print(f"Total records in source CSV: {len(df)}")
    print(f"Date range: {df['Date'].min()} to {df['Date'].max()}")
    print(f"Unique campaigns: {df['CampaignName'].nunique()}")
    
    # Check January 2025 data
    jan_data = df[(df['Date'] >= '2025-01-01') & (df['Date'] <= '2025-01-31')]
    print(f"\nJanuary 2025 records: {len(jan_data)}")
    print(f"January campaigns: {jan_data['CampaignName'].nunique()}")
    print(f"January spend total: ${jan_data['Cost'].sum():,.2f}")
    print(f"January sales total: ${jan_data['Sales'].sum():,.2f}")
    print(f"January clicks total: {jan_data['Clicks'].sum():,.0f}")
    print(f"January impressions total: {jan_data['Impressions'].sum():,.0f}")
    
    # Check February 2025 data
    feb_data = df[(df['Date'] >= '2025-02-01') & (df['Date'] <= '2025-02-28')]
    print(f"\nFebruary 2025 records: {len(feb_data)}")
    print(f"February campaigns: {feb_data['CampaignName'].nunique()}")
    print(f"February spend total: ${feb_data['Cost'].sum():,.2f}")
    print(f"February sales total: ${feb_data['Sales'].sum():,.2f}")
    print(f"February clicks total: {feb_data['Clicks'].sum():,.0f}")
    print(f"February impressions total: {feb_data['Impressions'].sum():,.0f}")
    
    # Check for data quality issues
    print(f"\nDATA QUALITY CHECKS:")
    print(f"Records with zero cost: {len(df[df['Cost'] == 0])}")
    print(f"Records with zero sales: {len(df[df['Sales'] == 0])}")
    print(f"Records with zero clicks: {len(df[df['Clicks'] == 0])}")
    print(f"Records with zero impressions: {len(df[df['Impressions'] == 0])}")
    
    # Check date coverage
    print(f"\nDATE COVERAGE:")
    date_counts = df.groupby(df['Date'].dt.date).size()
    print(f"Days with data: {len(date_counts)}")
    if len(date_counts) < 20:  # Show all if reasonable number
        for date, count in date_counts.items():
            print(f"  {date}: {count} records")
    
    return jan_data, feb_data

def examine_excel_data_scope():
    """Try to understand what data scope Excel is using"""
    excel_path = "../data/HydrapakUS_SPABridge_MoM_February 2025vsJanuary 2025.xlsx"
    
    print("\n\nEXCEL FILE ANALYSIS")
    print("=" * 80)
    
    wb = load_workbook(excel_path, data_only=True)
    print(f"Available worksheets: {wb.sheetnames}")
    
    # Check if there's a raw data sheet
    for sheet_name in wb.sheetnames:
        if 'data' in sheet_name.lower() or 'raw' in sheet_name.lower():
            print(f"\nFound potential data sheet: {sheet_name}")
            ws = wb[sheet_name]
            print(f"Rows: {ws.max_row}, Columns: {ws.max_column}")
    
    # Examine Campaign sheet structure
    ws = wb['Campaign']
    print(f"\nCampaign sheet structure:")
    print(f"Rows: {ws.max_row}, Columns: {ws.max_column}")
    
    # Look for any notes or metadata about data scope
    print(f"\nChecking for data scope information in first few rows:")
    for row in range(1, 15):
        first_cell = ws.cell(row=row, column=1).value
        if first_cell and isinstance(first_cell, str):
            if any(keyword in first_cell.lower() for keyword in ['total', 'data', 'period', 'filter', 'scope']):
                print(f"Row {row}: {first_cell}")

def compare_campaign_lists():
    """Compare which campaigns are included in each dataset"""
    print("\n\nCAMPAIGN COMPARISON")
    print("=" * 80)
    
    # Get CSV campaigns
    csv_path = "../data/Hydrapak YTD - campaign.csv"
    df = pd.read_csv(csv_path)
    df['Date'] = pd.to_datetime(df['DateKey'], format='%Y%m%d')
    
    jan_campaigns = set(df[(df['Date'] >= '2025-01-01') & (df['Date'] <= '2025-01-31')]['CampaignName'].unique())
    feb_campaigns = set(df[(df['Date'] >= '2025-02-01') & (df['Date'] <= '2025-02-28')]['CampaignName'].unique())
    csv_campaigns = jan_campaigns.union(feb_campaigns)
    
    print(f"CSV unique campaigns (Jan+Feb): {len(csv_campaigns)}")
    
    # Get Excel campaigns
    excel_path = "../data/HydrapakUS_SPABridge_MoM_February 2025vsJanuary 2025.xlsx"
    wb = load_workbook(excel_path, data_only=True)
    ws = wb['Campaign']
    
    excel_campaigns = set()
    for row in range(14, ws.max_row + 1):
        campaign_cell = ws.cell(row=row, column=1).value
        if campaign_cell and str(campaign_cell) != 'Total':
            excel_campaigns.add(str(campaign_cell).strip())
    
    print(f"Excel campaigns: {len(excel_campaigns)}")
    
    # Find differences
    csv_only = csv_campaigns - excel_campaigns
    excel_only = excel_campaigns - csv_campaigns
    common = csv_campaigns & excel_campaigns
    
    print(f"Common campaigns: {len(common)}")
    print(f"CSV only: {len(csv_only)}")
    print(f"Excel only: {len(excel_only)}")
    
    if len(csv_only) <= 10:
        print(f"\nCampaigns in CSV but not Excel:")
        for camp in sorted(csv_only):
            print(f"  - {camp}")
    
    if len(excel_only) <= 10:
        print(f"\nCampaigns in Excel but not CSV:")
        for camp in sorted(excel_only):
            print(f"  - {camp}")

if __name__ == "__main__":
    jan_data, feb_data = analyze_csv_source_data()
    examine_excel_data_scope()
    compare_campaign_lists()
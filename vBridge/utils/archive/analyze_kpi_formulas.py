import openpyxl
import pandas as pd
from openpyxl.utils import get_column_letter
import json

def analyze_excel_formulas(file_path):
    """Analyze formulas in Excel file tabs"""
    # Load the workbook
    wb = openpyxl.load_workbook(file_path, data_only=False)
    
    results = {}
    
    # Analyze p1_campaign_kpis and p2_campaign_kpis tabs
    for sheet_name in ['p1_campaign_kpis', 'p2_campaign_kpis']:
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            print(f"\n{'='*80}")
            print(f"Analyzing sheet: {sheet_name}")
            print('='*80)
            
            sheet_info = {
                'formulas': {},
                'headers': {},
                'dimensions': f"Rows: {ws.max_row}, Columns: {ws.max_column}"
            }
            
            # Get headers (assuming first row contains headers)
            headers = []
            for col in range(1, ws.max_column + 1):
                cell = ws.cell(row=1, column=col)
                if cell.value:
                    headers.append((col, cell.value))
                    sheet_info['headers'][get_column_letter(col)] = cell.value
            
            print(f"\nHeaders found:")
            for col, header in headers:
                print(f"  Column {get_column_letter(col)}: {header}")
            
            # Analyze formulas in all cells
            print(f"\nFormulas found:")
            formula_count = 0
            for row in range(1, min(ws.max_row + 1, 100)):  # Limit to first 100 rows for analysis
                for col in range(1, ws.max_column + 1):
                    cell = ws.cell(row=row, column=col)
                    if isinstance(cell.value, str) and cell.value.startswith('='):
                        formula_count += 1
                        cell_ref = f"{get_column_letter(col)}{row}"
                        sheet_info['formulas'][cell_ref] = {
                            'formula': cell.value,
                            'column_header': sheet_info['headers'].get(get_column_letter(col), 'No header')
                        }
                        print(f"\n  Cell {cell_ref} ({sheet_info['headers'].get(get_column_letter(col), 'No header')}):")
                        print(f"    Formula: {cell.value}")
            
            if formula_count == 0:
                print("  No formulas found in this sheet")
            
            results[sheet_name] = sheet_info
            
            # Also show a sample of data
            print(f"\nSample data (first 10 rows):")
            df = pd.read_excel(file_path, sheet_name=sheet_name, nrows=10)
            print(df.to_string())
        else:
            print(f"\nSheet '{sheet_name}' not found in workbook")
    
    # Save detailed results to JSON
    with open('../data/kpi_formula_analysis.json', 'w') as f:
        json.dump(results, f, indent=2)
    
    print(f"\n\nDetailed analysis saved to ../data/kpi_formula_analysis.json")
    
    # List all available sheets
    print(f"\n\nAll sheets in workbook:")
    for sheet in wb.sheetnames:
        print(f"  - {sheet}")
    
    return results

if __name__ == "__main__":
    file_path = "../data/HydrapakUS_SPABridge_MoM_February 2025vsJanuary 2025.xlsx"
    results = analyze_excel_formulas(file_path)
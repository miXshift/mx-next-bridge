#!/usr/bin/env python3
"""
Investigate why CSV contributions show 0 when Excel shows non-zero values
"""

import pandas as pd
from openpyxl import load_workbook

def load_excel_totals_row(excel_path):
    """Load the totals row from Excel"""
    wb = load_workbook(excel_path, data_only=True)
    ws = wb['Campaign']
    
    # Get headers from rows 12 and 13
    headers_row12 = []
    headers_row13 = []
    
    for col in range(1, ws.max_column + 1):
        val12 = ws.cell(row=12, column=col).value
        val13 = ws.cell(row=13, column=col).value
        headers_row12.append(val12 if val12 else '')
        headers_row13.append(val13 if val13 else '')
    
    # Create combined headers
    combined_headers = []
    for i, (h12, h13) in enumerate(zip(headers_row12, headers_row13)):
        if i == 0:
            combined_headers.append('Campaign Name')
        else:
            combined_headers.append(f"{h12} {h13}" if h12 and h13 else h13 if h13 else h12)
    
    # Find totals row (should be first data row after headers)
    totals_row = None
    for row in range(14, ws.max_row + 1):
        campaign_cell = ws.cell(row=row, column=1).value
        if campaign_cell and 'Total' in str(campaign_cell):
            totals_data = {}
            for col in range(1, ws.max_column + 1):
                cell_value = ws.cell(row=row, column=col).value
                header = combined_headers[col-1]
                totals_data[header] = cell_value
            return totals_data
    
    return None

def load_csv_totals_row(csv_path):
    """Load the totals row from CSV"""
    with open(csv_path, 'r') as f:
        lines = f.readlines()
    
    # Parse headers
    top_headers = lines[0].strip().split(',')
    sub_headers = lines[1].strip().split(',')
    
    # Create combined headers
    combined_headers = []
    for i, (top, sub) in enumerate(zip(top_headers, sub_headers)):
        if i == 0:
            combined_headers.append('Campaign Name')
        else:
            combined_headers.append(f"{top} {sub}")
    
    # First data row should be totals
    totals_line = lines[2].strip().split(',')
    if 'Total' in totals_line[0]:
        totals_data = {}
        for i, (header, value) in enumerate(zip(combined_headers, totals_line)):
            if i == 0:
                totals_data[header] = value
            else:
                try:
                    totals_data[header] = float(value) if value != '' else 0.0
                except ValueError:
                    totals_data[header] = value
        return totals_data
    
    return None

def examine_skyflask_contribution_calculation():
    """Examine the detailed contribution calculation for Skyflask campaign"""
    csv_path = "../output/period_comparison.csv"
    
    # Load CSV data
    with open(csv_path, 'r') as f:
        lines = f.readlines()
    
    # Parse headers
    top_headers = lines[0].strip().split(',')
    sub_headers = lines[1].strip().split(',')
    
    combined_headers = []
    for i, (top, sub) in enumerate(zip(top_headers, sub_headers)):
        if i == 0:
            combined_headers.append('Campaign Name')
        else:
            combined_headers.append(f"{top} {sub}")
    
    # Find totals and Skyflask rows
    totals_data = None
    skyflask_data = None
    
    for line in lines[2:]:
        if line.strip():
            row = line.strip().split(',')
            if 'Total' in row[0]:
                totals_data = {header: float(val) if val != '' and i > 0 else val 
                              for i, (header, val) in enumerate(zip(combined_headers, row))}
            elif '1000-CONQ-NONBRAND-ASIN-Skyflask' in row[0]:
                skyflask_data = {header: float(val) if val != '' and i > 0 else val 
                                for i, (header, val) in enumerate(zip(combined_headers, row))}
    
    if not totals_data or not skyflask_data:
        print("ERROR: Could not find totals or Skyflask data")
        return
    
    print("CONTRIBUTION CALCULATION INVESTIGATION")
    print("=" * 80)
    print("Formula: contribution = p1_mix * growth_rate * 10000")
    print("Where:")
    print("  p1_mix = campaign_jan_value / total_jan_value")
    print("  growth_rate = (campaign_feb_value / campaign_jan_value) - 1")
    print("=" * 80)
    
    # Check key metrics
    metrics_to_check = ['Spend', 'Total Ad Sales', 'ACoS', 'ROAS', 'Clicks']
    
    for metric in metrics_to_check:
        print(f"\n{metric.upper()} CONTRIBUTION CALCULATION:")
        print("-" * 50)
        
        # Get values
        jan_col = f"{metric} Jan 2025"
        feb_col = f"{metric} Feb 2025"
        contrib_col = f"{metric} Contribution"
        
        if jan_col in totals_data and jan_col in skyflask_data:
            total_jan = totals_data[jan_col]
            campaign_jan = skyflask_data[jan_col]
            campaign_feb = skyflask_data[feb_col]
            actual_contrib = skyflask_data[contrib_col]
            
            print(f"  Total January:     {total_jan:15.6f}")
            print(f"  Campaign January:  {campaign_jan:15.6f}")
            print(f"  Campaign February: {campaign_feb:15.6f}")
            
            if total_jan > 0 and campaign_jan > 0:
                p1_mix = campaign_jan / total_jan
                growth_rate = (campaign_feb / campaign_jan) - 1
                calculated_contrib = p1_mix * growth_rate * 10000
                
                print(f"  P1 Mix:            {p1_mix:15.10f}")
                print(f"  Growth Rate:       {growth_rate:15.10f}")
                print(f"  Calculated:        {calculated_contrib:15.6f}")
                print(f"  Actual CSV:        {actual_contrib:15.6f}")
                print(f"  Difference:        {abs(calculated_contrib - actual_contrib):15.6f}")
                
                if abs(calculated_contrib - actual_contrib) < 0.001:
                    print("  Status: ✓ MATCHES")
                else:
                    print("  Status: ✗ DIFFERS")
            else:
                print("  Status: Cannot calculate (zero values)")
        else:
            print(f"  ERROR: Missing columns for {metric}")

def compare_total_rows():
    """Compare total row values between Excel and CSV"""
    excel_path = "../data/HydrapakUS_SPABridge_MoM_February 2025vsJanuary 2025.xlsx"
    csv_path = "../output/period_comparison.csv"
    
    print("\nTOTAL ROW COMPARISON")
    print("=" * 80)
    
    excel_totals = load_excel_totals_row(excel_path)
    csv_totals = load_csv_totals_row(csv_path)
    
    if not excel_totals:
        print("ERROR: Could not load Excel totals")
        return
    
    if not csv_totals:
        print("ERROR: Could not load CSV totals")
        return
    
    # Compare key January totals (these are crucial for p1_mix calculation)
    key_metrics = ['Spend', 'Total Ad Sales', 'Clicks', 'Impressions']
    
    print("JANUARY TOTALS (crucial for p1_mix calculation):")
    print("-" * 60)
    print(f"{'Metric':<20} {'Excel':<15} {'CSV':<15} {'Difference':<15}")
    print("-" * 60)
    
    for metric in key_metrics:
        excel_col = None
        csv_col = f"{metric} Jan 2025"
        
        # Find Excel column
        for col in excel_totals.keys():
            if metric in col and 'January' in col:
                excel_col = col
                break
        
        if excel_col and csv_col in csv_totals:
            excel_val = float(excel_totals[excel_col]) if excel_totals[excel_col] is not None else 0
            csv_val = float(csv_totals[csv_col])
            diff = abs(excel_val - csv_val)
            
            status = "✓" if diff < 0.01 else "✗"
            print(f"{metric:<20} {excel_val:<15.2f} {csv_val:<15.2f} {diff:<15.6f} {status}")
        else:
            print(f"{metric:<20} {'NOT FOUND':<15} {'NOT FOUND':<15} {'N/A':<15}")

if __name__ == "__main__":
    examine_skyflask_contribution_calculation()
    compare_total_rows()
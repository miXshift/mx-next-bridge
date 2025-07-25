import pandas as pd
import openpyxl
from openpyxl import load_workbook

# Read the Excel file
excel_file = "../data/HydrapakUS_SPABridge_MoM_February 2025vsJanuary 2025.xlsx"

# Load workbook with formulas preserved
wb = load_workbook(excel_file, data_only=False, keep_vba=True)
ws = wb['Campaign']

print("=== DETAILED CAMPAIGN TAB ANALYSIS ===\n")

# Check for header information in the first rows
print("=== HEADER SECTION (Rows 1-12) ===\n")
for row in range(1, 13):
    row_values = []
    for col in range(1, min(11, ws.max_column + 1)):
        cell = ws.cell(row=row, column=col)
        if cell.value:
            row_values.append(f"{openpyxl.utils.get_column_letter(col)}: {cell.value}")
    if row_values:
        print(f"Row {row}: {' | '.join(row_values)}")

# Analyze the metric groups based on column headers
print("\n=== METRIC GROUPS ANALYSIS ===\n")
header_row = 13
metric_groups = []
current_group = []
group_start_col = 1

for col in range(1, ws.max_column + 1):
    cell = ws.cell(row=header_row, column=col)
    header = cell.value
    
    # Check the secondary header row (if exists)
    secondary_header = ws.cell(row=header_row-1, column=col).value
    
    if header:
        if col == 1:  # First column is Campaign
            continue
        elif "January 2025" in str(header) and current_group:
            # New metric group starts
            metric_groups.append({
                'start_col': group_start_col,
                'end_col': col - 1,
                'headers': current_group,
                'metric_name': secondary_header if secondary_header else f"Metric Group {len(metric_groups) + 1}"
            })
            current_group = [header]
            group_start_col = col
        else:
            current_group.append(header)

# Add the last group
if current_group:
    metric_groups.append({
        'start_col': group_start_col,
        'end_col': ws.max_column,
        'headers': current_group,
        'metric_name': ws.cell(row=header_row-1, column=group_start_col).value if ws.cell(row=header_row-1, column=group_start_col).value else f"Metric Group {len(metric_groups) + 1}"
    })

# Print metric groups
for i, group in enumerate(metric_groups):
    print(f"Metric Group {i+1}: Columns {openpyxl.utils.get_column_letter(group['start_col'])}-{openpyxl.utils.get_column_letter(group['end_col'])}")
    print(f"  Headers: {group['headers']}")
    print()

# Look for any formulas by checking raw cell formulas
print("=== FORMULA SEARCH (INCLUDING ARRAY FORMULAS) ===\n")
formula_count = 0
formula_examples = {}

for row in range(1, ws.max_row + 1):
    for col in range(1, ws.max_column + 1):
        cell = ws.cell(row=row, column=col)
        
        # Check if cell has a formula
        if hasattr(cell, '_value') and isinstance(cell._value, str) and cell._value.startswith('='):
            formula = cell._value
            col_letter = openpyxl.utils.get_column_letter(col)
            cell_ref = f"{col_letter}{row}"
            
            if col_letter not in formula_examples:
                formula_examples[col_letter] = []
            
            formula_examples[col_letter].append((cell_ref, formula))
            formula_count += 1
        
        # Also check for array formulas
        elif hasattr(cell, 'array_formula') and cell.array_formula:
            col_letter = openpyxl.utils.get_column_letter(col)
            cell_ref = f"{col_letter}{row}"
            print(f"Array formula at {cell_ref}: {cell.array_formula}")
            formula_count += 1

if formula_examples:
    print(f"Total formulas found: {formula_count}\n")
    for col in sorted(formula_examples.keys()):
        formulas = formula_examples[col]
        print(f"\nColumn {col} formulas:")
        # Show up to 3 examples
        for cell_ref, formula in formulas[:3]:
            print(f"  {cell_ref}: {formula}")
        if len(formulas) > 3:
            print(f"  ... and {len(formulas) - 3} more")
else:
    print("No formulas found. This appears to be a static data sheet.")

# Check for any hidden rows or columns
print("\n=== HIDDEN ROWS/COLUMNS CHECK ===\n")
hidden_rows = [row for row in range(1, ws.max_row + 1) if ws.row_dimensions[row].hidden]
hidden_cols = []
for col in range(1, ws.max_column + 1):
    col_letter = openpyxl.utils.get_column_letter(col)
    if col_letter in ws.column_dimensions and ws.column_dimensions[col_letter].hidden:
        hidden_cols.append(col_letter)

if hidden_rows:
    print(f"Hidden rows: {hidden_rows}")
else:
    print("No hidden rows found")

if hidden_cols:
    print(f"Hidden columns: {hidden_cols}")
else:
    print("No hidden columns found")

# Show sample data values with formatting
print("\n=== SAMPLE DATA WITH VALUES ===\n")
df = pd.read_excel(excel_file, sheet_name='Campaign', header=12)
print("First 3 campaign rows:")
for idx in range(min(3, len(df))):
    print(f"\nRow {idx+1} - Campaign: {df.iloc[idx, 0]}")
    for col_idx in range(1, min(6, len(df.columns))):
        value = df.iloc[idx, col_idx]
        col_name = df.columns[col_idx]
        print(f"  {col_name}: {value}")

# Check specific metric columns to understand the pattern
print("\n=== METRIC PATTERN ANALYSIS ===")
print("\nBased on column headers, the metrics appear to be:")
print("- Columns B-F: Metric 1 (January, February, Net Change, % Change, Contribution)")
print("- Columns G-K: Metric 2 (January, February, Net Change, % Change, Contribution)")
print("- Columns L-P: Metric 3 (January, February, Pts Change, % Change, Contribution)")
print("- And so on...")

# Try to identify what each metric represents by looking at row 12
print("\n=== METRIC IDENTIFIERS (Row 12) ===")
for col in range(2, ws.max_column + 1, 5):  # Check every 5th column
    metric_cell = ws.cell(row=12, column=col)
    if metric_cell.value:
        col_letter = openpyxl.utils.get_column_letter(col)
        print(f"Column {col_letter}: {metric_cell.value}")

wb.close()
#!/usr/bin/env python3
"""
Excel Formula Inspector - Extract actual formulas from Excel cells
Focuses on contribution calculation formulas for the Skyflask campaign
"""

from openpyxl import load_workbook

def inspect_excel_formulas():
    """Inspect Excel formulas for contribution calculations"""
    excel_path = "../data/HydrapakUS_SPABridge_MoM_February 2025vsJanuary 2025.xlsx"
    
    print("EXCEL FORMULA INSPECTION")
    print("="*60)
    print(f"File: {excel_path}")
    print(f"Target: 1000-CONQ-NONBRAND-ASIN-Skyflask contribution formulas")
    print("="*60)
    
    # Load with formulas (data_only=False)
    wb = load_workbook(excel_path, data_only=False)
    ws = wb['Campaign']
    
    # Find Skyflask row (we know it's row 33)
    skyflask_row = 33
    
    # First, let's map out the column structure
    print("\nCOLUMN STRUCTURE ANALYSIS:")
    print("-" * 40)
    
    kpi_row = 12  # Row with KPI names
    dimension_row = 13  # Row with period names
    
    kpi_columns = {}
    current_kpi = None
    
    for col in range(1, ws.max_column + 1):
        kpi_value = ws.cell(row=kpi_row, column=col).value
        dimension_value = ws.cell(row=dimension_row, column=col).value
        
        if kpi_value and str(kpi_value).strip():
            current_kpi = str(kpi_value).strip()
            kpi_columns[current_kpi] = []
        
        if current_kpi and dimension_value:
            kpi_columns[current_kpi].append({
                'column': col,
                'dimension': str(dimension_value).strip(),
                'header': f"{current_kpi} {dimension_value}"
            })
    
    # Focus on Spend and Total Ad Sales contribution columns
    target_kpis = ['Spend', 'Total Ad Sales']
    
    for kpi in target_kpis:
        if kpi in kpi_columns:
            print(f"\n{kpi.upper()} COLUMNS:")
            for col_info in kpi_columns[kpi]:
                print(f"  Col {col_info['column']:2d}: {col_info['dimension']}")
                
                if col_info['dimension'] == 'Contribution':
                    contribution_col = col_info['column']
                    
                    # Get the formula
                    formula_cell = ws.cell(row=skyflask_row, column=contribution_col)
                    formula = formula_cell.value if formula_cell.value else "No formula"
                    
                    # Get the calculated value  
                    wb_values = load_workbook(excel_path, data_only=True)
                    ws_values = wb_values['Campaign']
                    calculated_value = ws_values.cell(row=skyflask_row, column=contribution_col).value
                    
                    print(f"    📍 CONTRIBUTION ANALYSIS:")
                    print(f"       Formula: {formula}")
                    print(f"       Calculated Value: {calculated_value}")
                    print(f"       Cell Reference: {formula_cell.coordinate}")
    
    # Also check if there are any reference cells or constants
    print(f"\n" + "="*60)
    print("REFERENCE CELLS ANALYSIS")
    print("="*60)
    
    # Check for any cells that might contain constants or reference values
    print("\nChecking for potential reference values in first few rows:")
    for row in range(1, 15):
        for col in range(1, min(10, ws.max_column + 1)):
            cell_value = ws.cell(row=row, column=col).value
            if isinstance(cell_value, (int, float)) and abs(cell_value) > 1000:
                print(f"  Row {row}, Col {col}: {cell_value}")
    
    # Let's also check the totals row for reference
    total_row = 14
    print(f"\nTOTALS ROW ANALYSIS (Row {total_row}):")
    
    for kpi in ['Spend', 'Total Ad Sales']:
        if kpi in kpi_columns:
            for col_info in kpi_columns[kpi]:
                if col_info['dimension'] in ['January 2025', 'February 2025']:
                    col = col_info['column']
                    total_value = ws.cell(row=total_row, column=col).value
                    print(f"  {kpi} {col_info['dimension']}: {total_value}")

def verify_contribution_calculations():
    """Verify our Python calculations step by step"""
    print(f"\n" + "="*60)
    print("PYTHON CALCULATION VERIFICATION")
    print("="*60)
    
    # Skyflask data (from our analysis)
    skyflask_data = {
        'spend_jan': 290.39,
        'spend_feb': 235.04,
        'sales_jan': 1244.60,
        'sales_feb': 1208.56
    }
    
    # Total data (from Excel totals row)
    total_data = {
        'spend_jan': 53057.41,  # This might be wrong - need to verify
        'sales_jan': 255793.53  # This might be wrong - need to verify  
    }
    
    print("\nSPEND CONTRIBUTION CALCULATION:")
    print("-" * 30)
    
    p1_mix_spend = skyflask_data['spend_jan'] / total_data['spend_jan']
    growth_rate_spend = (skyflask_data['spend_feb'] / skyflask_data['spend_jan']) - 1
    contribution_spend = p1_mix_spend * growth_rate_spend * 10000
    
    print(f"Campaign Jan Spend: ${skyflask_data['spend_jan']}")
    print(f"Campaign Feb Spend: ${skyflask_data['spend_feb']}")
    print(f"Total Jan Spend: ${total_data['spend_jan']}")
    print(f"P1_Mix: {p1_mix_spend:.8f}")
    print(f"Growth Rate: {growth_rate_spend:.8f}")
    print(f"Contribution: {contribution_spend:.4f} basis points")
    print(f"Rounded Contribution: {round(contribution_spend)}")
    
    print("\nTOTAL AD SALES CONTRIBUTION CALCULATION:")
    print("-" * 30)
    
    p1_mix_sales = skyflask_data['sales_jan'] / total_data['sales_jan']
    growth_rate_sales = (skyflask_data['sales_feb'] / skyflask_data['sales_jan']) - 1
    contribution_sales = p1_mix_sales * growth_rate_sales * 10000
    
    print(f"Campaign Jan Sales: ${skyflask_data['sales_jan']}")
    print(f"Campaign Feb Sales: ${skyflask_data['sales_feb']}")
    print(f"Total Jan Sales: ${total_data['sales_jan']}")
    print(f"P1_Mix: {p1_mix_sales:.8f}")
    print(f"Growth Rate: {growth_rate_sales:.8f}")
    print(f"Contribution: {contribution_sales:.4f} basis points")
    print(f"Rounded Contribution: {round(contribution_sales)}")

if __name__ == "__main__":
    inspect_excel_formulas()
    verify_contribution_calculations()
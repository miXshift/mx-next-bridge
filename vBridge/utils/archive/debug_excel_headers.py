#!/usr/bin/env python3
"""
Debug script to examine Excel headers and find the Skyflask campaign data
"""

from openpyxl import load_workbook

def debug_excel_structure():
    excel_path = "../data/HydrapakUS_SPABridge_MoM_February 2025vsJanuary 2025.xlsx"
    
    wb = load_workbook(excel_path, data_only=True)
    ws = wb['Campaign']
    
    print("EXCEL CAMPAIGN TAB STRUCTURE")
    print("="*60)
    print(f"Total rows: {ws.max_row}, Total columns: {ws.max_column}")
    
    # Look for headers by examining rows with multiple non-empty columns
    print("\nLooking for rows with multiple columns (likely header rows):")
    for row in range(1, min(30, ws.max_row + 1)):
        non_empty_cols = []
        for col in range(1, ws.max_column + 1):
            cell_value = ws.cell(row=row, column=col).value
            if cell_value:
                non_empty_cols.append((col, str(cell_value)[:50]))  # Truncate long values
        
        if len(non_empty_cols) > 5:  # Rows with many columns are likely headers
            print(f"\nRow {row} has {len(non_empty_cols)} non-empty columns:")
            for col_num, value in non_empty_cols[:15]:  # Show first 15
                print(f"  Col {col_num}: '{value}'")
            if len(non_empty_cols) > 15:
                print(f"  ... and {len(non_empty_cols) - 15} more columns")
    
    # Find Skyflask row
    skyflask_row = None
    for row in range(2, ws.max_row + 1):
        campaign_cell = ws.cell(row=row, column=1).value
        if campaign_cell and "Skyflask" in str(campaign_cell):
            skyflask_row = row
            print(f"\nFound Skyflask at row {row}: '{campaign_cell}'")
            break
    
    # If we found data rows, look at the row just before the first data row for headers
    if skyflask_row:
        potential_header_row = skyflask_row - 1
        print(f"\nChecking potential header row {potential_header_row}:")
        headers = []
        for col in range(1, ws.max_column + 1):
            header = ws.cell(row=potential_header_row, column=col).value
            if header:
                headers.append((col, str(header)))
        
        if headers:
            print(f"Found {len(headers)} headers in row {potential_header_row}:")
            for col_num, header in headers:
                print(f"  Col {col_num}: '{header}'")
        
        # Show Skyflask data with these headers
        print(f"\nSkyflask campaign data with identified headers:")
        for col in range(1, min(ws.max_column + 1, 20)):  # Show first 20 columns
            header = ws.cell(row=potential_header_row, column=col).value
            value = ws.cell(row=skyflask_row, column=col).value
            if header or value:
                print(f"  Col {col} - {header}: {value}")
    
    # Look specifically for contribution-related keywords in all cells
    print(f"\nSearching for contribution-related terms across the entire sheet:")
    contribution_keywords = ['Contribution', 'Contrib', 'Revenue', 'Spend', 'Cost', 'ROAS', 'CPA', 'Jan', 'Feb']
    
    found_terms = {}
    for row in range(1, min(30, ws.max_row + 1)):
        for col in range(1, ws.max_column + 1):
            cell_value = ws.cell(row=row, column=col).value
            if cell_value:
                cell_str = str(cell_value)
                for keyword in contribution_keywords:
                    if keyword.lower() in cell_str.lower():
                        if keyword not in found_terms:
                            found_terms[keyword] = []
                        found_terms[keyword].append(f"Row {row}, Col {col}: '{cell_str}'")
    
    for keyword, locations in found_terms.items():
        print(f"\nFound '{keyword}' in:")
        for location in locations[:5]:  # Show first 5 occurrences
            print(f"  {location}")
        if len(locations) > 5:
            print(f"  ... and {len(locations) - 5} more locations")

if __name__ == "__main__":
    debug_excel_structure()
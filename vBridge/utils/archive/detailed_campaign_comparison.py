#!/usr/bin/env python3
"""
Detailed KPI comparison for specific campaign: 1000-CONQ-NONBRAND-ASIN-Skyflask
Compares every metric between Excel and CSV output
"""

import pandas as pd
import numpy as np
from openpyxl import load_workbook

def load_excel_campaign_row(excel_path, campaign_name):
    """Load specific campaign row from Excel with proper two-tier headers"""
    wb = load_workbook(excel_path, data_only=True)
    ws = wb['Campaign']
    
    # Get two-tier headers from rows 12 and 13
    headers_row12 = []
    headers_row13 = []
    
    for col in range(1, ws.max_column + 1):
        val12 = ws.cell(row=12, column=col).value
        val13 = ws.cell(row=13, column=col).value
        headers_row12.append(val12 if val12 else '')
        headers_row13.append(val13 if val13 else '')
    
    # Create combined headers
    combined_headers = []
    for i, (h12, h13) in enumerate(zip(headers_row12, headers_row13)):
        if i == 0:
            combined_headers.append('Campaign Name')
        else:
            combined_headers.append(f"{h12} {h13}" if h12 and h13 else h13 if h13 else h12)
    
    # Find campaign row (data starts from row 14)
    for row in range(14, ws.max_row + 1):
        campaign_cell = ws.cell(row=row, column=1).value
        if campaign_cell and campaign_name in str(campaign_cell):
            campaign_data = {}
            for col in range(1, ws.max_column + 1):
                cell_value = ws.cell(row=row, column=col).value
                header = combined_headers[col-1]
                campaign_data[header] = cell_value
            return campaign_data
    
    return None

def load_csv_campaign_row(csv_path, campaign_name):
    """Load specific campaign row from CSV"""
    with open(csv_path, 'r') as f:
        lines = f.readlines()
    
    # Parse headers
    top_headers = lines[0].strip().split(',')
    sub_headers = lines[1].strip().split(',')
    
    # Create combined headers
    combined_headers = []
    for i, (top, sub) in enumerate(zip(top_headers, sub_headers)):
        if i == 0:
            combined_headers.append('Campaign Name')
        else:
            combined_headers.append(f"{top} {sub}")
    
    # Find campaign row
    for line in lines[2:]:
        if line.strip():
            row = line.strip().split(',')
            if campaign_name in row[0]:
                campaign_data = {}
                for i, (header, value) in enumerate(zip(combined_headers, row)):
                    if i == 0:
                        campaign_data[header] = value
                    else:
                        try:
                            campaign_data[header] = float(value) if value != '' else 0.0
                        except ValueError:
                            campaign_data[header] = value
                return campaign_data
    
    return None

def generate_detailed_comparison_report():
    """Generate detailed comparison for 1000-CONQ-NONBRAND-ASIN-Skyflask"""
    campaign_name = "1000-CONQ-NONBRAND-ASIN-Skyflask"
    excel_path = "../data/HydrapakUS_SPABridge_MoM_February 2025vsJanuary 2025.xlsx"
    csv_path = "../output/period_comparison.csv"
    
    print("DETAILED CAMPAIGN KPI COMPARISON")
    print("="*80)
    print(f"Campaign: {campaign_name}")
    print(f"Excel Source: {excel_path}")
    print(f"CSV Source: {csv_path}")
    print("="*80)
    
    # Load campaign data
    excel_data = load_excel_campaign_row(excel_path, campaign_name)
    csv_data = load_csv_campaign_row(csv_path, campaign_name)
    
    if not excel_data:
        print("ERROR: Campaign not found in Excel data")
        return
    
    if not csv_data:
        print("ERROR: Campaign not found in CSV data")
        return
    
    print("✓ Campaign found in both datasets\n")
    
    # Define KPI mappings and expected metrics
    kpi_mappings = [
        # Spend
        ("Spend", "Jan 2025", "Spend Jan 2025"),
        ("Spend", "Feb 2025", "Spend Feb 2025"),
        ("Spend", "Net Change", "Spend Net Change"),
        ("Spend", "% Change", "Spend % Change"),
        ("Spend", "Contribution", "Spend Contribution"),
        
        # Total Ad Sales
        ("Total Ad Sales", "Jan 2025", "Total Ad Sales Jan 2025"),
        ("Total Ad Sales", "Feb 2025", "Total Ad Sales Feb 2025"),
        ("Total Ad Sales", "Net Change", "Total Ad Sales Net Change"),
        ("Total Ad Sales", "% Change", "Total Ad Sales % Change"),
        ("Total Ad Sales", "Contribution", "Total Ad Sales Contribution"),
        
        # ACoS
        ("ACoS", "Jan 2025", "ACoS Jan 2025"),
        ("ACoS", "Feb 2025", "ACoS Feb 2025"),
        ("ACoS", "Pts Change", "ACoS Pts Change"),
        ("ACoS", "% Change", "ACoS % Change"),
        ("ACoS", "Contribution", "ACoS Contribution"),
        
        # ROAS
        ("ROAS", "Jan 2025", "ROAS Jan 2025"),
        ("ROAS", "Feb 2025", "ROAS Feb 2025"),
        ("ROAS", "Net Change", "ROAS Net Change"),
        ("ROAS", "% Change", "ROAS % Change"),
        ("ROAS", "Contribution", "ROAS Contribution"),
        
        # Conversion Rate
        ("Conversion Rate", "Jan 2025", "Conversion Rate Jan 2025"),
        ("Conversion Rate", "Feb 2025", "Conversion Rate Feb 2025"),
        ("Conversion Rate", "Pts Change", "Conversion Rate Pts Change"),
        ("Conversion Rate", "% Change", "Conversion Rate % Change"),
        ("Conversion Rate", "Contribution", "Conversion Rate Contribution"),
        
        # Impressions
        ("Impressions", "Jan 2025", "Impressions Jan 2025"),
        ("Impressions", "Feb 2025", "Impressions Feb 2025"),
        ("Impressions", "Net Change", "Impressions Net Change"),
        ("Impressions", "% Change", "Impressions % Change"),
        ("Impressions", "Contribution", "Impressions Contribution"),
        
        # Clicks
        ("Clicks", "Jan 2025", "Clicks Jan 2025"),
        ("Clicks", "Feb 2025", "Clicks Feb 2025"),
        ("Clicks", "Net Change", "Clicks Net Change"),
        ("Clicks", "% Change", "Clicks % Change"),
        ("Clicks", "Contribution", "Clicks Contribution"),
        
        # CTR
        ("CTR", "Jan 2025", "CTR Jan 2025"),
        ("CTR", "Feb 2025", "CTR Feb 2025"),
        ("CTR", "Pts Change", "CTR Pts Change"),
        ("CTR", "% Change", "CTR % Change"),
        ("CTR", "Contribution", "CTR Contribution"),
        
        # CPC
        ("CPC", "Jan 2025", "CPC Jan 2025"),
        ("CPC", "Feb 2025", "CPC Feb 2025"),
        ("CPC", "Net Change", "CPC Net Change"),
        ("CPC", "% Change", "CPC % Change"),
        ("CPC", "Contribution", "CPC Contribution"),
        
        # Same SKU Ad Sales
        ("Same SKU Ad Sales", "Jan 2025", "Same SKU Ad Sales Jan 2025"),
        ("Same SKU Ad Sales", "Feb 2025", "Same SKU Ad Sales Feb 2025"),
        ("Same SKU Ad Sales", "Net Change", "Same SKU Ad Sales Net Change"),
        ("Same SKU Ad Sales", "% Change", "Same SKU Ad Sales % Change"),
        ("Same SKU Ad Sales", "Contribution", "Same SKU Ad Sales Contribution"),
        
        # Other SKU Sales
        ("Other SKU Sales", "Jan 2025", "Other SKU Sales Jan 2025"),
        ("Other SKU Sales", "Feb 2025", "Other SKU Sales Feb 2025"),
        ("Other SKU Sales", "Net Change", "Other SKU Sales Net Change"),
        ("Other SKU Sales", "% Change", "Other SKU Sales % Change"),
        ("Other SKU Sales", "Contribution", "Other SKU Sales Contribution"),
        
        # Same SKU Ad Orders
        ("Same SKU Ad Orders", "Jan 2025", "Same SKU Ad Orders Jan 2025"),
        ("Same SKU Ad Orders", "Feb 2025", "Same SKU Ad Orders Feb 2025"),
        ("Same SKU Ad Orders", "Net Change", "Same SKU Ad Orders Net Change"),
        ("Same SKU Ad Orders", "% Change", "Same SKU Ad Orders % Change"),
        ("Same SKU Ad Orders", "Contribution", "Same SKU Ad Orders Contribution"),
        
        # Other SKU Ad Orders
        ("Other SKU Ad Orders", "Jan 2025", "Other SKU Ad Orders Jan 2025"),
        ("Other SKU Ad Orders", "Feb 2025", "Other SKU Ad Orders Feb 2025"),
        ("Other SKU Ad Orders", "Net Change", "Other SKU Ad Orders Net Change"),
        ("Other SKU Ad Orders", "% Change", "Other SKU Ad Orders % Change"),
        ("Other SKU Ad Orders", "Contribution", "Other SKU Ad Orders Contribution"),
        
        # Total Ad Orders
        ("Total Ad Orders", "Jan 2025", "Total Ad Orders Jan 2025"),
        ("Total Ad Orders", "Feb 2025", "Total Ad Orders Feb 2025"),
        ("Total Ad Orders", "Net Change", "Total Ad Orders Net Change"),
        ("Total Ad Orders", "% Change", "Total Ad Orders % Change"),
        ("Total Ad Orders", "Contribution", "Total Ad Orders Contribution"),
    ]
    
    # Compare each KPI
    total_comparisons = 0
    perfect_matches = 0
    close_matches = 0
    significant_differences = 0
    
    current_kpi = None
    
    for kpi, dimension, csv_col in kpi_mappings:
        if kpi != current_kpi:
            if current_kpi is not None:
                print()
            print(f"{kpi.upper()} METRICS:")
            print("-" * 50)
            current_kpi = kpi
        
        # Find matching Excel column with flexible matching
        excel_col = None
        # Try exact match first
        for col in excel_data.keys():
            if kpi in str(col) and dimension in str(col):
                excel_col = col
                break
        
        # If not found, try alternative dimension names
        if not excel_col:
            alt_dimension = dimension
            if dimension == "Jan 2025":
                alt_dimension = "January 2025"
            elif dimension == "Feb 2025":
                alt_dimension = "February 2025"
            elif dimension == "Net Change":
                alt_dimension = "Net Change"
            elif dimension == "Pts Change":
                alt_dimension = "Pts Change"
            elif dimension == "% Change":
                alt_dimension = "% Change"
            elif dimension == "Contribution":
                alt_dimension = "Contribution"
            
            for col in excel_data.keys():
                if kpi in str(col) and alt_dimension in str(col):
                    excel_col = col
                    break
        
        if not excel_col:
            print(f"  {dimension:15}: Excel column not found")
            continue
        
        # Get values
        excel_val = excel_data[excel_col]
        csv_val = csv_data.get(csv_col, 'NOT_FOUND')
        
        # Convert to numeric if possible
        try:
            excel_num = float(excel_val) if excel_val is not None else 0.0
            csv_num = float(csv_val) if csv_val != 'NOT_FOUND' else None
            
            # Handle percentage format differences for rate metrics
            if kpi in ['ACoS', 'CTR', 'Conversion Rate'] and excel_num > 1:
                # Excel stores percentages (23.33), CSV stores decimals (0.233)
                excel_num = excel_num / 100
                
        except (ValueError, TypeError):
            excel_num = excel_val
            csv_num = csv_val
        
        # Compare (handle decimal percentage conversion)
        total_comparisons += 1
        
        if csv_num is None:
            status = "✗ CSV_MISSING"
            print(f"  {dimension:15}: Excel={excel_num:>12}, CSV=NOT_FOUND, {status}")
        elif isinstance(excel_num, (int, float)) and isinstance(csv_num, (int, float)):
            # Convert Excel percentage to decimal for % Change comparisons
            if "% Change" in dimension:
                excel_decimal = excel_num / 100  # Convert Excel percentage to decimal
                diff = abs(excel_decimal - csv_num)
                rel_diff = (diff / abs(excel_decimal) * 100) if excel_decimal != 0 else 0
                
                if diff < 0.0001:
                    status = "✓ PERFECT"
                    perfect_matches += 1
                elif diff < 0.001:
                    status = "✓ CLOSE"
                    close_matches += 1
                else:
                    status = "✗ DIFF"
                    significant_differences += 1
                
                print(f"  {dimension:15}: Excel={excel_decimal:>12.6f}, CSV={csv_num:>12.6f}, Diff={diff:>8.6f} ({rel_diff:>5.1f}%) {status}")
            else:
                # Regular comparison for non-percentage values
                diff = abs(excel_num - csv_num)
                rel_diff = (diff / abs(excel_num) * 100) if excel_num != 0 else 0
                
                if diff < 0.001:
                    status = "✓ PERFECT"
                    perfect_matches += 1
                elif diff < 0.1:
                    status = "✓ CLOSE"
                    close_matches += 1
                else:
                    status = "✗ DIFF"
                    significant_differences += 1
                
                print(f"  {dimension:15}: Excel={excel_num:>12.6f}, CSV={csv_num:>12.6f}, Diff={diff:>8.6f} ({rel_diff:>5.1f}%) {status}")
        else:
            if str(excel_num).strip() == str(csv_num).strip():
                status = "✓ MATCH"
                perfect_matches += 1
            else:
                status = "✗ DIFF"
                significant_differences += 1
            print(f"  {dimension:15}: Excel='{excel_num}', CSV='{csv_num}', {status}")
    
    # Summary
    print("\n" + "="*80)
    print("COMPARISON SUMMARY")
    print("="*80)
    print(f"Total comparisons: {total_comparisons}")
    if total_comparisons > 0:
        print(f"Perfect matches:   {perfect_matches} ({perfect_matches/total_comparisons*100:.1f}%)")
        print(f"Close matches:     {close_matches} ({close_matches/total_comparisons*100:.1f}%)")
        print(f"Differences:       {significant_differences} ({significant_differences/total_comparisons*100:.1f}%)")
    else:
        print("No valid comparisons found - check column mapping")
    
    if total_comparisons > 0:
        accuracy = (perfect_matches + close_matches) / total_comparisons * 100
        print(f"\nOverall Accuracy: {accuracy:.1f}%")
    else:
        accuracy = 0
    
    if accuracy >= 95:
        print("✅ EXCELLENT - Output matches Excel calculations very closely")
    elif accuracy >= 90:
        print("✅ GOOD - Output matches Excel calculations well")
    elif accuracy >= 80:
        print("⚠️  FAIR - Some differences found, review needed")
    else:
        print("❌ POOR - Significant differences found, investigation required")

if __name__ == "__main__":
    generate_detailed_comparison_report()
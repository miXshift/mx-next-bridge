#!/usr/bin/env python3
"""
Updated precise comparison that properly handles Excel column structure
"""

import pandas as pd
from openpyxl import load_workbook

def load_excel_campaign_data(excel_path, campaign_name):
    """Load Excel data with proper column handling"""
    wb = load_workbook(excel_path, data_only=True)
    ws = wb['Campaign']
    
    # Get the two-tier headers
    headers_row12 = []
    headers_row13 = []
    
    for col in range(1, ws.max_column + 1):
        val12 = ws.cell(row=12, column=col).value
        val13 = ws.cell(row=13, column=col).value
        headers_row12.append(val12 if val12 else '')
        headers_row13.append(val13 if val13 else '')
    
    # Create combined headers
    combined_headers = []
    for i, (h12, h13) in enumerate(zip(headers_row12, headers_row13)):
        if i == 0:
            combined_headers.append('Campaign Name')
        else:
            combined_headers.append(f"{h12} {h13}" if h12 and h13 else h13 if h13 else h12)
    
    # Find campaign row (data starts from row 14)
    for row in range(14, ws.max_row + 1):
        campaign_cell = ws.cell(row=row, column=1).value
        if campaign_cell and campaign_name in str(campaign_cell):
            campaign_data = {}
            for col in range(1, ws.max_column + 1):
                cell_value = ws.cell(row=row, column=col).value
                header = combined_headers[col-1]
                campaign_data[header] = cell_value
            return campaign_data
    
    return None

def load_csv_campaign_row(csv_path, campaign_name):
    """Load specific campaign row from CSV"""
    with open(csv_path, 'r') as f:
        lines = f.readlines()
    
    # Parse headers
    top_headers = lines[0].strip().split(',')
    sub_headers = lines[1].strip().split(',')
    
    # Create combined headers
    combined_headers = []
    for i, (top, sub) in enumerate(zip(top_headers, sub_headers)):
        if i == 0:
            combined_headers.append('Campaign Name')
        else:
            combined_headers.append(f"{top} {sub}")
    
    # Find campaign row
    for line in lines[2:]:
        if line.strip():
            row = line.strip().split(',')
            if campaign_name in row[0]:
                campaign_data = {}
                for i, (header, value) in enumerate(zip(combined_headers, row)):
                    if i == 0:
                        campaign_data[header] = value
                    else:
                        try:
                            campaign_data[header] = float(value) if value != '' else 0.0
                        except ValueError:
                            campaign_data[header] = value
                return campaign_data
    
    return None

def compare_with_high_precision():
    """Compare with high precision"""
    campaign_name = "Skyflask"  # Search for campaigns containing "Skyflask"
    excel_path = "../data/HydrapakUS_SPABridge_MoM_February 2025vsJanuary 2025.xlsx"
    csv_path = "../output/period_comparison.csv"
    
    print("HIGH PRECISION COMPARISON")
    print("=" * 80)
    
    # Load data
    excel_data = load_excel_campaign_data(excel_path, campaign_name)
    csv_data = load_csv_campaign_row(csv_path, campaign_name)
    
    if not excel_data:
        print("ERROR: Campaign not found in Excel data")
        return
    
    if not csv_data:
        print("ERROR: Campaign not found in CSV data")
        return
    
    print(f"Excel Campaign: {[k for k in excel_data.keys() if 'Campaign' in k]}")
    print(f"CSV Campaign: {csv_data.get('Campaign Name', 'NOT_FOUND')}")
    print()
    
    # Compare contribution values
    metrics = ['Spend', 'Total Ad Sales', 'ACoS', 'ROAS', 'Conversion Rate',
               'Impressions', 'Clicks', 'CTR', 'CPC', 'Same SKU Ad Sales',
               'Other SKU Sales', 'Same SKU Ad Orders', 'Other SKU Ad Orders', 'Total Ad Orders']
    
    print("CONTRIBUTION COMPARISON:")
    print("=" * 120)
    print(f"{'Metric':<25} {'Excel Value':<20} {'CSV Value':<20} {'Difference':<20} {'Status'}")
    print("-" * 120)
    
    significant_diffs = []
    
    for metric in metrics:
        # Find Excel contribution column
        excel_contrib_col = None
        for col in excel_data.keys():
            if metric in col and 'Contribution' in col:
                excel_contrib_col = col
                break
        
        # CSV contribution column
        csv_contrib_col = f"{metric} Contribution"
        
        if excel_contrib_col and csv_contrib_col in csv_data:
            excel_val = float(excel_data[excel_contrib_col]) if excel_data[excel_contrib_col] is not None else 0.0
            csv_val = float(csv_data[csv_contrib_col]) if csv_data[csv_contrib_col] is not None else 0.0
            
            diff = abs(excel_val - csv_val)
            
            if diff < 0.01:
                status = "✓ MATCH"
            elif diff < 1.0:
                status = "⚠ MINOR"
            else:
                status = "✗ MAJOR"
                significant_diffs.append((metric, excel_val, csv_val, diff))
            
            print(f"{metric:<25} {excel_val:<20.10f} {csv_val:<20.10f} {diff:<20.10f} {status}")
        else:
            print(f"{metric:<25} {'NOT FOUND':<20} {'NOT FOUND':<20} {'N/A':<20} {'✗ MISSING'}")
    
    # Detailed analysis of major differences
    if significant_diffs:
        print("\n" + "="*80)
        print("DETAILED ANALYSIS OF MAJOR DIFFERENCES")
        print("="*80)
        
        for metric, excel_val, csv_val, diff in significant_diffs:
            print(f"\n{metric} CONTRIBUTION:")
            print(f"  Excel: {excel_val:15.10f}")
            print(f"  CSV:   {csv_val:15.10f}")
            print(f"  Diff:  {diff:15.10f}")
            print(f"  Ratio: {csv_val/excel_val if excel_val != 0 else 'N/A'}")
    
    # Also check percentage changes
    print("\n" + "="*80)
    print("PERCENTAGE CHANGE COMPARISON")
    print("="*80)
    
    for metric in metrics:
        excel_pct_col = None
        for col in excel_data.keys():
            if metric in col and '% Change' in col:
                excel_pct_col = col
                break
        
        csv_pct_col = f"{metric} % Change"
        
        if excel_pct_col and csv_pct_col in csv_data:
            excel_val = float(excel_data[excel_pct_col]) if excel_data[excel_pct_col] is not None else 0.0
            csv_val = float(csv_data[csv_pct_col]) if csv_data[csv_pct_col] is not None else 0.0
            
            diff = abs(excel_val - csv_val)
            if diff > 0.01:  # Only show meaningful differences
                status = "✓ CLOSE" if diff < 0.1 else "✗ DIFF"
                print(f"{metric:<25} Excel: {excel_val:10.6f}, CSV: {csv_val:10.6f}, Diff: {diff:8.6f} {status}")

if __name__ == "__main__":
    compare_with_high_precision()
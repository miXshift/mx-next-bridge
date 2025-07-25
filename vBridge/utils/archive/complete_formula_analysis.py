import openpyxl
import pandas as pd
from openpyxl.utils import get_column_letter
import json

def complete_formula_analysis(file_path):
    """Complete analysis of formulas in KPI tabs"""
    wb = openpyxl.load_workbook(file_path, data_only=False)
    
    print("\n" + "="*100)
    print("COMPLETE FORMULA ANALYSIS FOR KPI TABS")
    print("="*100)
    
    # First, analyze the Campaign sheet to understand VLOOKUP references
    if 'Campaign' in wb.sheetnames:
        print("\nCAMPAIGN SHEET STRUCTURE (Reference for VLOOKUPs):")
        print("-"*50)
        ws = wb['Campaign']
        print(f"Dimensions: {ws.max_row} rows x {ws.max_column} columns")
        print("\nFirst 6 columns (used in VLOOKUP formulas):")
        for col in range(1, 7):
            cell = ws.cell(row=1, column=col)
            print(f"  Column {get_column_letter(col)}: {cell.value}")
    
    # Analyze each KPI sheet
    for sheet_name in ['p1_campaign_kpis', 'p2_campaign_kpis']:
        if sheet_name not in wb.sheetnames:
            continue
            
        ws = wb[sheet_name]
        print(f"\n\n{'='*100}")
        print(f"SHEET: {sheet_name}")
        print('='*100)
        
        # Find header row
        header_row = None
        for row in range(1, min(10, ws.max_row + 1)):
            cell_value = ws.cell(row=row, column=1).value
            if cell_value and 'Campaign' in str(cell_value):
                header_row = row
                break
        
        print(f"\nSHEET STRUCTURE:")
        print(f"  Total dimensions: {ws.max_row} rows x {ws.max_column} columns")
        print(f"  Header row: {header_row}")
        
        # Get all headers
        headers = {}
        if header_row:
            print(f"\nCOLUMN HEADERS:")
            for col in range(1, ws.max_column + 1):
                cell = ws.cell(row=header_row, column=col)
                if cell.value:
                    col_letter = get_column_letter(col)
                    headers[col_letter] = str(cell.value)
                    print(f"  {col_letter}: {cell.value}")
        
        # Analyze formulas by type
        print(f"\nFORMULA ANALYSIS:")
        
        # 1. Aggregation formulas (typically in row 1 or 2)
        print("\n1. AGGREGATION FORMULAS (Summary rows):")
        for row in range(1, min(5, ws.max_row + 1)):
            for col in range(1, ws.max_column + 1):
                cell = ws.cell(row=row, column=col)
                if cell.value and isinstance(cell.value, str) and cell.value.startswith('='):
                    if any(func in cell.value.upper() for func in ['SUM', 'AVERAGE', 'COUNT']):
                        col_letter = get_column_letter(col)
                        print(f"  Cell {col_letter}{row}: {cell.value}")
                        print(f"    Purpose: Aggregates {headers.get(col_letter, 'Unknown')} data")
        
        # 2. VLOOKUP formulas
        print("\n2. VLOOKUP FORMULAS (Data mapping from Campaign sheet):")
        vlookup_found = False
        for row in range(1, min(ws.max_row + 1, 100)):
            for col in range(1, ws.max_column + 1):
                cell = ws.cell(row=row, column=col)
                if cell.value and isinstance(cell.value, str) and 'VLOOKUP' in cell.value.upper():
                    if not vlookup_found:  # Show only first example of each pattern
                        col_letter = get_column_letter(col)
                        print(f"  Example in {col_letter}{row}: {cell.value}")
                        # Parse VLOOKUP parameters
                        if 'Campaign!$A:$F,3,0' in cell.value:
                            print(f"    Maps: Campaign Name (column A) → Column C from Campaign sheet")
                        vlookup_found = True
                        break
        
        # 3. Calculation formulas
        print("\n3. CALCULATION FORMULAS:")
        calc_patterns = {}
        for row in range(1, min(ws.max_row + 1, 100)):
            for col in range(1, ws.max_column + 1):
                cell = ws.cell(row=row, column=col)
                if cell.value and isinstance(cell.value, str) and cell.value.startswith('='):
                    if not any(func in cell.value.upper() for func in ['SUM', 'VLOOKUP', 'AVERAGE', 'COUNT']):
                        col_letter = get_column_letter(col)
                        # Generalize pattern
                        import re
                        pattern = re.sub(r'\d+', 'N', cell.value)
                        if pattern not in calc_patterns:
                            calc_patterns[pattern] = {
                                'example': cell.value,
                                'cell': f"{col_letter}{row}",
                                'column': col_letter
                            }
        
        for pattern, info in calc_patterns.items():
            print(f"  Pattern: {pattern}")
            print(f"    Example: {info['example']} in {info['cell']}")
            if '=' in pattern and '-' in pattern:
                print(f"    Purpose: Calculates difference/variance")
        
        # 4. Analyze which columns have no formulas (direct data)
        print("\n4. DIRECT DATA COLUMNS (no formulas):")
        formula_columns = set()
        for row in range(1, ws.max_row + 1):
            for col in range(1, ws.max_column + 1):
                cell = ws.cell(row=row, column=col)
                if cell.value and isinstance(cell.value, str) and cell.value.startswith('='):
                    formula_columns.add(get_column_letter(col))
        
        for col_letter, header in headers.items():
            if col_letter not in formula_columns:
                print(f"  {col_letter}: {header}")
        
        # Show data flow
        print(f"\n5. DATA FLOW SUMMARY:")
        if sheet_name == 'p1_campaign_kpis':
            print("  - This sheet appears to contain DIRECT campaign performance data")
            print("  - No formulas found in data rows - all values are imported directly")
            print("  - Metrics include: Spend, Sales, ACoS, ROAS, CTR, CPC, Orders, etc.")
        elif sheet_name == 'p2_campaign_kpis':
            print("  - Row 1: Contains SUM formulas to aggregate totals")
            print("  - Column C: Uses VLOOKUP to fetch values from Campaign sheet (column 3)")
            print("  - Column D: Calculates variance (Column C - Column B)")
            print("  - Other columns: Direct data import for metrics")

if __name__ == "__main__":
    file_path = "../data/HydrapakUS_SPABridge_MoM_February 2025vsJanuary 2025.xlsx"
    complete_formula_analysis(file_path)
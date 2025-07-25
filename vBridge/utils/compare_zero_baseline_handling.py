#!/usr/bin/env python3
"""
Compare how Excel handles zero P1 (baseline) values versus our implementation
"""

import pandas as pd
from openpyxl import load_workbook
import sys
import os

# Add the src directory to the path
sys.path.append(os.path.join(os.path.dirname(__file__), '..', 'src'))

from campaign_bridge import CampaignBridge

def analyze_excel_zero_handling():
    """Analyze how Excel handles campaigns with zero baseline values"""
    
    print("EXCEL ZERO BASELINE HANDLING ANALYSIS")
    print("=" * 80)
    
    # Load Excel file
    excel_path = "data/HydrapakUS_SPABridge_MoM_February 2025vsJanuary 2025.xlsx"
    wb = load_workbook(excel_path, data_only=True)
    ws = wb['Campaign']
    
    # Find data rows (starting from row 15 based on previous analysis)
    data_start_row = 15
    
    # Column mapping for Spend
    spend_columns = {
        'jan': None,
        'feb': None,
        'contrib': None
    }
    
    # Find Spend columns
    for col in range(1, ws.max_column + 1):
        kpi_cell = ws.cell(row=12, column=col).value
        dim_cell = ws.cell(row=13, column=col).value
        
        if kpi_cell and 'Spend' in str(kpi_cell):
            if dim_cell:
                if 'January' in str(dim_cell):
                    spend_columns['jan'] = col
                elif 'February' in str(dim_cell):
                    spend_columns['feb'] = col
                elif 'Contribution' in str(dim_cell):
                    spend_columns['contrib'] = col
    
    print(f"Found Spend columns:")
    print(f"  January: Column {spend_columns['jan']}")
    print(f"  February: Column {spend_columns['feb']}")
    print(f"  Contribution: Column {spend_columns['contrib']}")
    
    # Find campaigns with zero January spend
    zero_baseline_campaigns = []
    
    row = data_start_row
    while row <= ws.max_row:
        campaign_name = ws.cell(row=row, column=1).value
        if not campaign_name or 'Total' in str(campaign_name):
            row += 1
            continue
            
        jan_spend = ws.cell(row=row, column=spend_columns['jan']).value
        feb_spend = ws.cell(row=row, column=spend_columns['feb']).value
        contrib = ws.cell(row=row, column=spend_columns['contrib']).value
        
        # Check if January is zero/None but February has value
        if (jan_spend is None or jan_spend == 0) and feb_spend and feb_spend > 0:
            zero_baseline_campaigns.append({
                'row': row,
                'campaign': campaign_name,
                'jan_spend': jan_spend or 0,
                'feb_spend': feb_spend,
                'contribution': contrib or 0
            })
        
        row += 1
    
    print(f"\nFound {len(zero_baseline_campaigns)} campaigns with zero January spend in Excel:")
    print("-" * 80)
    
    for camp in zero_baseline_campaigns[:10]:  # Show first 10
        print(f"Row {camp['row']:3}: {camp['campaign'][:40]:<40}")
        print(f"         Jan: ${camp['jan_spend']:8.2f}, Feb: ${camp['feb_spend']:8.2f}, Contrib: {camp['contribution']:8.2f}")
        
        # Check the formula in this cell
        wb_formulas = load_workbook(excel_path, data_only=False)
        ws_formulas = wb_formulas['Campaign']
        formula_cell = ws_formulas.cell(row=camp['row'], column=spend_columns['contrib'])
        formula = formula_cell.value
        if formula and isinstance(formula, str) and formula.startswith('='):
            print(f"         Formula: {formula[:60]}...")
        print()
    
    return zero_baseline_campaigns

def analyze_csv_zero_handling():
    """Analyze how our implementation handles campaigns with zero baseline values"""
    
    print("\nCSV/PYTHON ZERO BASELINE HANDLING ANALYSIS")
    print("=" * 80)
    
    # Generate fresh data
    bridge = CampaignBridge('data/Hydrapak YTD - campaign.csv')
    bridge.load_data()
    bridge_df = bridge.calculate_bridge()
    
    # Find campaigns with zero January spend
    zero_baseline_campaigns = []
    
    # Skip total row (index 0)
    for idx, row in bridge_df.iloc[1:].iterrows():
        jan_spend = row['Spend - January 2025']
        feb_spend = row['Spend - February 2025']
        contribution = row['Spend - Contribution']
        
        if jan_spend == 0 and feb_spend > 0:
            zero_baseline_campaigns.append({
                'campaign': row['Campaign'],
                'jan_spend': jan_spend,
                'feb_spend': feb_spend,
                'contribution': contribution
            })
    
    print(f"Found {len(zero_baseline_campaigns)} campaigns with zero January spend in CSV:")
    print("-" * 80)
    
    for camp in zero_baseline_campaigns[:10]:  # Show first 10
        print(f"{camp['campaign'][:50]:<50}")
        print(f"  Jan: ${camp['jan_spend']:8.2f}, Feb: ${camp['feb_spend']:8.2f}, Contrib: {camp['contribution']:8.2f}")
        
        # Calculate what contribution should be with dummy value
        if bridge_df.iloc[0]['Spend - January 2025'] > 0:
            # Our implementation uses dummy value 0.0000001
            dummy_jan = 0.0000001
            p1_mix = dummy_jan / bridge_df.iloc[0]['Spend - January 2025']
            growth_rate = (camp['feb_spend'] / dummy_jan) - 1
            expected_contrib = p1_mix * growth_rate * 10000
            print(f"  Expected with dummy (0.0000001): {expected_contrib:.6f}")
        print()
    
    return zero_baseline_campaigns

def compare_handling():
    """Compare Excel vs CSV handling of zero baseline"""
    
    print("\nCOMPARISON SUMMARY")
    print("=" * 80)
    
    excel_zeros = analyze_excel_zero_handling()
    csv_zeros = analyze_csv_zero_handling()
    
    print(f"\nZERO BASELINE CAMPAIGNS:")
    print(f"Excel count: {len(excel_zeros)}")
    print(f"CSV count:   {len(csv_zeros)}")
    
    print("\nKEY DIFFERENCES:")
    print("-" * 60)
    print("1. Excel Approach:")
    print("   - Appears to exclude zero baseline campaigns from contribution")
    print("   - Or uses a different formula/handling for these cases")
    print("   - Maintains mathematical consistency")
    
    print("\n2. CSV/Python Approach:")
    print("   - Uses dummy value (0.0000001) for zero baseline")
    print("   - Allows contribution calculation but breaks summation")
    print("   - Creates mismatch between individual sum and total")
    
    print("\n3. Mathematical Impact:")
    print("   - Excel: Sum of individual contributions = Total contribution")
    print("   - CSV: Sum of individual contributions ≠ Total contribution")
    print("   - Difference is due to zero baseline handling")

if __name__ == "__main__":
    compare_handling()
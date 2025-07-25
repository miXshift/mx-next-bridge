import pandas as pd
import openpyxl
from openpyxl import load_workbook

# Read the Excel file
excel_file = "../data/HydrapakUS_SPABridge_MoM_February 2025vsJanuary 2025.xlsx"

# Load workbook to check for formulas
wb = load_workbook(excel_file, data_only=False)

# Check if 'Campaign' sheet exists
if 'Campaign' in wb.sheetnames:
    ws = wb['Campaign']
    
    print("=== CAMPAIGN TAB ANALYSIS ===\n")
    
    # Get dimensions
    print(f"Sheet dimensions: {ws.max_row} rows x {ws.max_column} columns\n")
    
    # Get column headers (assuming they're in row 1)
    print("Column headers:")
    headers = []
    for col in range(1, ws.max_column + 1):
        cell = ws.cell(row=1, column=col)
        if cell.value:
            headers.append((col, cell.value))
            print(f"  Column {col}: {cell.value}")
    
    print("\n=== FORMULAS FOUND ===\n")
    
    # Search for formulas in the sheet
    formula_count = 0
    for row in range(1, min(ws.max_row + 1, 100)):  # Check first 100 rows for formulas
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row=row, column=col)
            if isinstance(cell.value, str) and cell.value.startswith('='):
                formula_count += 1
                col_letter = openpyxl.utils.get_column_letter(col)
                print(f"Cell {col_letter}{row}: {cell.value}")
                if formula_count > 50:  # Limit output
                    print("... (showing first 50 formulas)")
                    break
        if formula_count > 50:
            break
    
    if formula_count == 0:
        print("No formulas found in the first 100 rows")
    
    # Check formatting of specific cells
    print("\n=== FORMATTING EXAMPLES ===\n")
    
    # Check a few cells for number formatting
    for row in [2, 3, 4]:  # Check rows 2, 3, 4
        for col in range(1, min(ws.max_column + 1, 10)):  # First 10 columns
            cell = ws.cell(row=row, column=col)
            if cell.number_format and cell.number_format != 'General':
                col_letter = openpyxl.utils.get_column_letter(col)
                print(f"Cell {col_letter}{row}: Format = {cell.number_format}")
    
    # Also read with pandas to see the data structure
    print("\n=== DATA PREVIEW (using pandas) ===\n")
    df = pd.read_excel(excel_file, sheet_name='Campaign')
    print(f"DataFrame shape: {df.shape}")
    print("\nFirst 5 rows:")
    print(df.head())
    
    print("\nColumn data types:")
    print(df.dtypes)
    
else:
    print("'campaign' sheet not found in the Excel file")
    print("Available sheets:", wb.sheetnames)

wb.close()
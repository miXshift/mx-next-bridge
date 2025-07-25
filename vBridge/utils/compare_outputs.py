#!/usr/bin/env python3
"""
Compare period_comparison.csv output with Excel campaign tab
Validates that our Python output matches the original Excel calculations
"""

import pandas as pd
import numpy as np
import openpyxl
from openpyxl import load_workbook

def load_excel_campaign_data(excel_path):
    """Load campaign data from Excel file"""
    print("Loading Excel campaign data...")
    
    # Load with openpyxl to preserve formatting
    wb = load_workbook(excel_path, data_only=True)
    ws = wb['Campaign']
    
    # Convert to DataFrame
    data = []
    headers = []
    
    # Get headers from first row
    for col in range(1, ws.max_column + 1):
        cell_value = ws.cell(row=1, column=col).value
        headers.append(cell_value if cell_value else f"Col_{col}")
    
    # Get data rows
    for row in range(2, ws.max_row + 1):
        row_data = []
        for col in range(1, ws.max_column + 1):
            cell_value = ws.cell(row=row, column=col).value
            row_data.append(cell_value)
        data.append(row_data)
    
    excel_df = pd.DataFrame(data, columns=headers)
    print(f"Excel data loaded: {len(excel_df)} rows, {len(excel_df.columns)} columns")
    
    return excel_df

def load_csv_data(csv_path):
    """Load our generated CSV data with two-tier headers"""
    print("Loading CSV data...")
    
    # Read the CSV manually to handle two-tier headers
    with open(csv_path, 'r') as f:
        lines = f.readlines()
    
    # Parse headers
    top_headers = lines[0].strip().split(',')
    sub_headers = lines[1].strip().split(',')
    
    print(f"Top headers count: {len(top_headers)}")
    print(f"Sub headers count: {len(sub_headers)}")
    
    # Create combined headers for comparison
    combined_headers = []
    for i, (top, sub) in enumerate(zip(top_headers, sub_headers)):
        if i == 0:  # Campaign Name column
            combined_headers.append('Campaign Name')
        else:
            if sub in ['Net Change', 'Pts Change']:
                combined_headers.append(f"{top} {sub}")
            else:
                combined_headers.append(f"{top} {sub}")
    
    # Parse data rows
    data = []
    for line in lines[2:]:
        if line.strip():
            row = line.strip().split(',')
            # Debug only first few rows
            if len(data) < 3:
                print(f"Row {len(data)+1} length: {len(row)}, Headers length: {len(combined_headers)}")
            if len(row) != len(combined_headers):
                print(f"WARNING: Row length mismatch. Adjusting...")
                # Pad or trim row to match headers
                while len(row) < len(combined_headers):
                    row.append('')
                row = row[:len(combined_headers)]
            
            # Convert numeric values
            converted_row = [row[0]]  # Keep campaign name as string
            for val in row[1:]:
                try:
                    converted_row.append(float(val) if val != '' else 0.0)
                except ValueError:
                    converted_row.append(val)
            data.append(converted_row)
    
    csv_df = pd.DataFrame(data, columns=combined_headers)
    print(f"CSV data loaded: {len(csv_df)} rows, {len(csv_df.columns)} columns")
    
    return csv_df, top_headers, sub_headers

def compare_campaign_names(excel_df, csv_df):
    """Compare campaign names between Excel and CSV"""
    print("\n" + "="*60)
    print("CAMPAIGN NAME COMPARISON")
    print("="*60)
    
    excel_campaigns = set(excel_df.iloc[:, 0].astype(str).str.strip())
    csv_campaigns = set(csv_df.iloc[:, 0].astype(str).str.strip())
    
    print(f"Excel campaigns: {len(excel_campaigns)}")
    print(f"CSV campaigns: {len(csv_campaigns)}")
    
    missing_in_csv = excel_campaigns - csv_campaigns
    missing_in_excel = csv_campaigns - excel_campaigns
    
    if missing_in_csv:
        print(f"\nCampaigns in Excel but missing in CSV ({len(missing_in_csv)}):")
        for campaign in sorted(missing_in_csv):
            print(f"  - {campaign}")
    
    if missing_in_excel:
        print(f"\nCampaigns in CSV but missing in Excel ({len(missing_in_excel)}):")
        for campaign in sorted(missing_in_excel):
            print(f"  - {campaign}")
    
    common_campaigns = excel_campaigns & csv_campaigns
    print(f"\nCommon campaigns: {len(common_campaigns)}")
    
    return common_campaigns

def compare_totals_row(excel_df, csv_df):
    """Compare the totals row between Excel and CSV"""
    print("\n" + "="*60)
    print("TOTALS ROW COMPARISON")
    print("="*60)
    
    # Find totals row in each dataset
    excel_total_row = excel_df[excel_df.iloc[:, 0].astype(str).str.contains('Total', case=False, na=False)]
    csv_total_row = csv_df[csv_df.iloc[:, 0].astype(str).str.contains('Total', case=False, na=False)]
    
    if excel_total_row.empty:
        print("ERROR: No totals row found in Excel data")
        return
    
    if csv_total_row.empty:
        print("ERROR: No totals row found in CSV data")
        return
    
    print("Comparing key metrics in totals row...")
    
    # Define key metrics to compare (using patterns that should exist in both)
    metrics_to_compare = [
        ('Spend', 'Jan 2025'),
        ('Spend', 'Feb 2025'), 
        ('Total Ad Sales', 'Jan 2025'),
        ('Total Ad Sales', 'Feb 2025'),
        ('ACoS', 'Jan 2025'),
        ('ACoS', 'Feb 2025'),
        ('ROAS', 'Jan 2025'),
        ('ROAS', 'Feb 2025')
    ]
    
    differences_found = False
    
    for metric, period in metrics_to_compare:
        # Try to find matching columns
        csv_col = f"{metric} {period}"
        excel_cols = [col for col in excel_df.columns if metric in str(col) and period in str(col)]
        
        if csv_col in csv_df.columns and excel_cols:
            excel_col = excel_cols[0]
            csv_val = csv_total_row[csv_col].iloc[0]
            excel_val = excel_total_row[excel_col].iloc[0]
            
            try:
                csv_val = float(csv_val) if csv_val is not None else 0
                excel_val = float(excel_val) if excel_val is not None else 0
                
                diff = abs(csv_val - excel_val)
                rel_diff = (diff / excel_val * 100) if excel_val != 0 else 0
                
                status = "✓ MATCH" if diff < 0.01 else "✗ DIFF"
                print(f"{metric} {period:12}: Excel={excel_val:12.2f}, CSV={csv_val:12.2f}, Diff={diff:8.2f} ({rel_diff:5.1f}%) {status}")
                
                if diff >= 0.01:
                    differences_found = True
            except (ValueError, TypeError):
                print(f"{metric} {period:12}: Could not compare values (Excel: {excel_val}, CSV: {csv_val})")
    
    if not differences_found:
        print("\n✓ All compared totals match within tolerance!")
    else:
        print("\n✗ Some differences found in totals row")

def compare_sample_campaigns(excel_df, csv_df, common_campaigns, sample_size=5):
    """Compare a sample of campaigns between Excel and CSV"""
    print("\n" + "="*60)
    print(f"SAMPLE CAMPAIGN COMPARISON ({sample_size} campaigns)")
    print("="*60)
    
    # Take a sample of common campaigns (excluding Total)
    sample_campaigns = list(common_campaigns)[:sample_size]
    if 'Total' in sample_campaigns:
        sample_campaigns.remove('Total')
        if len(common_campaigns) > sample_size:
            sample_campaigns.append(list(common_campaigns)[sample_size])
    
    for campaign in sample_campaigns:
        print(f"\nCampaign: {campaign}")
        print("-" * 40)
        
        excel_row = excel_df[excel_df.iloc[:, 0].astype(str).str.strip() == campaign]
        csv_row = csv_df[csv_df.iloc[:, 0].astype(str).str.strip() == campaign]
        
        if excel_row.empty or csv_row.empty:
            print(f"  Campaign not found in both datasets")
            continue
        
        # Compare Spend Jan 2025
        try:
            excel_spend_cols = [col for col in excel_df.columns if 'Spend' in str(col) and 'Jan' in str(col)]
            csv_spend_col = 'Spend Jan 2025'
            
            if excel_spend_cols and csv_spend_col in csv_df.columns:
                excel_val = float(excel_row[excel_spend_cols[0]].iloc[0]) if excel_row[excel_spend_cols[0]].iloc[0] is not None else 0
                csv_val = float(csv_row[csv_spend_col].iloc[0]) if csv_row[csv_spend_col].iloc[0] is not None else 0
                diff = abs(excel_val - csv_val)
                status = "✓" if diff < 0.01 else "✗"
                print(f"  Spend Jan:  Excel={excel_val:8.2f}, CSV={csv_val:8.2f}, Diff={diff:6.2f} {status}")
        except Exception as e:
            print(f"  Spend Jan:  Error comparing - {e}")

def generate_comparison_report():
    """Generate a comprehensive comparison report"""
    excel_path = "../data/HydrapakUS_SPABridge_MoM_February 2025vsJanuary 2025.xlsx"
    csv_path = "../output/period_comparison.csv"
    
    print("CAMPAIGN BRIDGE OUTPUT COMPARISON REPORT")
    print("="*80)
    print(f"Excel file: {excel_path}")
    print(f"CSV file: {csv_path}")
    print("="*80)
    
    try:
        # Load data
        excel_df = load_excel_campaign_data(excel_path)
        csv_df, top_headers, sub_headers = load_csv_data(csv_path)
        
        # Compare campaign names
        common_campaigns = compare_campaign_names(excel_df, csv_df)
        
        # Compare totals row
        compare_totals_row(excel_df, csv_df)
        
        # Compare sample campaigns
        compare_sample_campaigns(excel_df, csv_df, common_campaigns)
        
        # Summary
        print("\n" + "="*60)
        print("SUMMARY")
        print("="*60)
        print(f"✓ Data loaded successfully from both files")
        print(f"✓ Campaign comparison completed")
        print(f"✓ {len(common_campaigns)} campaigns found in both datasets")
        print("\nFor detailed validation, review the comparison results above.")
        
    except Exception as e:
        print(f"ERROR: {e}")
        import traceback
        traceback.print_exc()

if __name__ == "__main__":
    generate_comparison_report()
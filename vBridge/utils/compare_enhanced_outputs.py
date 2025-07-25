#!/usr/bin/env python3
"""
Standalone Enhanced Output Comparison Tool
Compares enhanced MixBridge CSV outputs with Excel source of truth
"""

import pandas as pd
import openpyxl
import numpy as np
import json
from pathlib import Path
from datetime import datetime
import sys
import os

def load_excel_source(excel_path: str) -> pd.DataFrame:
    """Load Excel source of truth with proper header detection"""
    print(f"📊 Loading Excel source: {excel_path}")
    
    try:
        wb = openpyxl.load_workbook(excel_path, data_only=True)
        
        # Find Campaign sheet
        if 'Campaign' in wb.sheetnames:
            ws = wb['Campaign']
        elif 'Campaign Tab' in wb.sheetnames:
            ws = wb['Campaign Tab']
        else:
            ws = wb.active
            
        print(f"✅ Using sheet: {ws.title}")
        
        # Find header row by looking for 'Campaign' in first column
        header_row = 13  # Default based on analysis
        for row in range(1, 20):
            cell_value = ws.cell(row=row, column=1).value
            if cell_value and str(cell_value).strip().lower() == 'campaign':
                header_row = row
                break
        
        print(f"✅ Header row: {header_row}")
        
        # Extract headers
        headers = []
        for col in range(1, ws.max_column + 1):
            cell_value = ws.cell(row=header_row, column=col).value
            headers.append(cell_value if cell_value else f'Col_{col}')
        
        # Extract data
        data = []
        for row in range(header_row + 1, ws.max_row + 1):
            row_data = []
            for col in range(1, ws.max_column + 1):
                cell_value = ws.cell(row=row, column=col).value
                row_data.append(cell_value)
            if any(val is not None for val in row_data):
                data.append(row_data)
        
        excel_df = pd.DataFrame(data, columns=headers)
        excel_df = excel_df.dropna(how='all')
        
        print(f"✅ Excel data loaded: {len(excel_df)} rows, {len(excel_df.columns)} columns")
        return excel_df
        
    except Exception as e:
        print(f"❌ Error loading Excel: {e}")
        return None

def load_csv_output(csv_path: str) -> pd.DataFrame:
    """Load enhanced CSV output"""
    print(f"📈 Loading CSV: {csv_path}")
    
    try:
        csv_df = pd.read_csv(csv_path)
        print(f"✅ CSV data loaded: {len(csv_df)} rows, {len(csv_df.columns)} columns")
        return csv_df
        
    except Exception as e:
        print(f"❌ Error loading CSV: {e}")
        return None

def find_latest_csv(csv_directory: str = 'output/analyses') -> str:
    """Find the latest enhanced CSV file"""
    analyses_dir = Path(csv_directory)
    
    if not analyses_dir.exists():
        print(f"❌ Analyses directory not found: {analyses_dir}")
        return None
    
    # Find mixbridge CSV files
    csv_files = list(analyses_dir.glob('mixbridge_*.csv'))
    
    if not csv_files:
        print(f"❌ No mixbridge CSV files found in {analyses_dir}")
        return None
    
    # Sort by modification time, get latest
    latest_file = max(csv_files, key=lambda x: x.stat().st_mtime)
    print(f"✅ Latest CSV found: {latest_file.name}")
    
    return str(latest_file)

def compare_campaign_coverage(excel_df: pd.DataFrame, csv_df: pd.DataFrame) -> dict:
    """Compare campaign coverage between Excel and CSV"""
    print("\n🔍 CAMPAIGN COVERAGE ANALYSIS")
    print("-" * 50)
    
    # Extract campaign lists
    excel_campaigns = set(excel_df['Campaign'].dropna().astype(str))
    csv_campaigns = set(csv_df['Campaign'].dropna().astype(str))
    
    # Clean campaign names
    excel_campaigns = {c.strip() for c in excel_campaigns if c.strip()}
    csv_campaigns = {c.strip() for c in csv_campaigns if c.strip()}
    
    common = excel_campaigns.intersection(csv_campaigns)
    excel_only = excel_campaigns - csv_campaigns
    csv_only = csv_campaigns - excel_campaigns
    
    coverage_rate = (len(common) / max(len(excel_campaigns), len(csv_campaigns))) * 100
    
    print(f"📊 Excel campaigns: {len(excel_campaigns)}")
    print(f"📊 CSV campaigns: {len(csv_campaigns)}")
    print(f"✅ Common campaigns: {len(common)}")
    print(f"⚠️  Excel only: {len(excel_only)}")
    print(f"⚠️  CSV only: {len(csv_only)}")
    print(f"📊 Coverage rate: {coverage_rate:.1f}%")
    
    if excel_only:
        print(f"📋 Excel-only campaigns (first 5): {list(excel_only)[:5]}")
    if csv_only:
        print(f"📋 CSV-only campaigns (first 5): {list(csv_only)[:5]}")
    
    return {
        'excel_campaigns': len(excel_campaigns),
        'csv_campaigns': len(csv_campaigns),
        'common_campaigns': len(common),
        'coverage_rate': coverage_rate,
        'excel_only': len(excel_only),
        'csv_only': len(csv_only)
    }

def compare_total_metrics(excel_df: pd.DataFrame, csv_df: pd.DataFrame) -> dict:
    """Compare total row metrics for accuracy assessment"""
    print("\n🎯 TOTAL ROW ACCURACY ANALYSIS")
    print("-" * 50)
    
    # Find total rows
    excel_total = excel_df[
        excel_df['Campaign'].astype(str).str.contains('total', case=False, na=False)
    ]
    csv_total = csv_df[
        csv_df['Campaign'].astype(str).str.contains('total', case=False, na=False)
    ]
    
    if len(excel_total) == 0 or len(csv_total) == 0:
        print("❌ Total rows not found in one or both files")
        return {"error": "Total rows not found"}
    
    print(f"✅ Excel Total rows: {len(excel_total)}")
    print(f"✅ CSV Total rows: {len(csv_total)}")
    
    # Key metric comparisons
    comparisons = [
        ('January 2025', 'Spend - January 2025', 'January Spend'),
        ('February 2025', 'Spend - February 2025', 'February Spend'),
        ('Net Change', 'Spend - Net Change', 'Net Change'),
        ('% Change', 'Spend - % Change', 'Percent Change')
    ]
    
    results = []
    accuracy_scores = []
    
    print(f"\n📊 Key Metrics Comparison:")
    
    for excel_col, csv_col, metric_name in comparisons:
        if excel_col in excel_total.columns and csv_col in csv_total.columns:
            try:
                excel_val = float(excel_total.iloc[0][excel_col])
                csv_val = float(csv_total.iloc[0][csv_col])
                
                diff = abs(csv_val - excel_val)
                rel_diff = (diff / abs(excel_val) * 100) if excel_val != 0 else 0
                
                # Determine accuracy grade
                if rel_diff < 0.01:
                    grade = 'PERFECT'
                    grade_emoji = '🎯'
                elif rel_diff < 0.1:
                    grade = 'EXCELLENT'
                    grade_emoji = '✅'
                elif rel_diff < 1.0:
                    grade = 'GOOD'
                    grade_emoji = '✅'
                elif rel_diff < 5.0:
                    grade = 'FAIR'
                    grade_emoji = '⚠️'
                else:
                    grade = 'POOR'
                    grade_emoji = '❌'
                
                results.append({
                    'metric': metric_name,
                    'excel_value': excel_val,
                    'csv_value': csv_val,
                    'difference': diff,
                    'relative_difference': rel_diff,
                    'grade': grade
                })
                
                accuracy_scores.append(rel_diff)
                
                print(f"  📈 {metric_name}:")
                print(f"      Excel: ${excel_val:,.2f}")
                print(f"      CSV:   ${csv_val:,.2f}")
                print(f"      Grade: {grade_emoji} {grade} ({rel_diff:.4f}% error)")
                
            except Exception as e:
                print(f"  ❌ Error comparing {metric_name}: {e}")
    
    # Overall assessment
    if accuracy_scores:
        avg_error = np.mean(accuracy_scores)
        max_error = max(accuracy_scores)
        accuracy_score = 100 - avg_error
        
        if avg_error < 0.01:
            overall_grade = 'PERFECT'
            grade_emoji = '🎯'
        elif avg_error < 0.1:
            overall_grade = 'EXCELLENT'
            grade_emoji = '✅'
        elif avg_error < 1.0:
            overall_grade = 'GOOD'
            grade_emoji = '✅'
        elif avg_error < 5.0:
            overall_grade = 'FAIR'
            grade_emoji = '⚠️'
        else:
            overall_grade = 'POOR'
            grade_emoji = '❌'
        
        print(f"\n🏆 OVERALL ACCURACY ASSESSMENT:")
        print(f"    Grade: {grade_emoji} {overall_grade}")
        print(f"    Accuracy Score: {accuracy_score:.2f}%")
        print(f"    Average Error: {avg_error:.4f}%")
        print(f"    Maximum Error: {max_error:.4f}%")
        
        return {
            'metric_results': results,
            'overall_grade': overall_grade,
            'accuracy_score': accuracy_score,
            'average_error': avg_error,
            'maximum_error': max_error
        }
    
    return {"error": "No metrics successfully compared"}

def compare_sample_campaigns(excel_df: pd.DataFrame, csv_df: pd.DataFrame, sample_size: int = 5) -> dict:
    """Compare sample campaigns for detailed verification"""
    print(f"\n🔍 SAMPLE CAMPAIGN VERIFICATION ({sample_size} campaigns)")
    print("-" * 50)
    
    # Find common campaigns (excluding total)
    excel_campaigns = set(excel_df['Campaign'].dropna().astype(str))
    csv_campaigns = set(csv_df['Campaign'].dropna().astype(str))
    
    # Remove 'Total' and empty strings
    excel_campaigns = {c.strip() for c in excel_campaigns 
                      if c.strip() and 'total' not in c.lower()}
    csv_campaigns = {c.strip() for c in csv_campaigns 
                    if c.strip() and 'total' not in c.lower()}
    
    common = list(excel_campaigns.intersection(csv_campaigns))
    
    if len(common) < sample_size:
        sample_size = len(common)
        print(f"⚠️  Only {sample_size} common campaigns available")
    
    if sample_size == 0:
        print("❌ No common campaigns found for sampling")
        return {"error": "No common campaigns found"}
    
    # Take first N campaigns
    sample_campaigns = common[:sample_size]
    results = []
    
    for campaign in sample_campaigns:
        excel_row = excel_df[excel_df['Campaign'] == campaign]
        csv_row = csv_df[csv_df['Campaign'] == campaign]
        
        if len(excel_row) > 0 and len(csv_row) > 0:
            try:
                excel_spend = float(excel_row.iloc[0]['January 2025'])
                csv_spend = float(csv_row.iloc[0]['Spend - January 2025'])
                diff = abs(csv_spend - excel_spend)
                rel_diff = (diff / abs(excel_spend) * 100) if excel_spend != 0 else 0
                
                status = '✅' if rel_diff < 1.0 else '⚠️'
                
                results.append({
                    'campaign': campaign,
                    'excel_value': excel_spend,
                    'csv_value': csv_spend,
                    'relative_difference': rel_diff,
                    'status': 'PASS' if rel_diff < 1.0 else 'REVIEW'
                })
                
                # Truncate long campaign names for display
                display_name = campaign[:50] + "..." if len(campaign) > 50 else campaign
                print(f"  {status} {display_name}: {rel_diff:.2f}% error")
                
            except Exception as e:
                print(f"  ❌ Error comparing {campaign}: {e}")
    
    pass_count = sum(1 for r in results if r['status'] == 'PASS')
    pass_rate = (pass_count / len(results) * 100) if results else 0
    
    print(f"\n📊 Sample Results: {pass_count}/{len(results)} passed ({pass_rate:.1f}%)")
    
    return {
        'sample_size': len(results),
        'pass_count': pass_count,
        'pass_rate': pass_rate,
        'campaign_results': results
    }

def save_comparison_results(results: dict, output_dir: str = 'output/reports'):
    """Save comparison results to JSON file"""
    output_path = Path(output_dir)
    output_path.mkdir(parents=True, exist_ok=True)
    
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    filename = f"comparison_results_{timestamp}.json"
    filepath = output_path / filename
    
    with open(filepath, 'w') as f:
        json.dump(results, f, indent=2, default=str)
    
    print(f"\n💾 Results saved: {filepath}")

def main():
    """Main comparison execution"""
    print("🎯 ENHANCED MIXBRIDGE COMPARISON TOOL")
    print("=" * 80)
    
    # Configuration
    excel_path = 'data/HydrapakUS_SPABridge_MoM_February 2025vsJanuary 2025.xlsx'
    
    # Handle command line arguments
    if len(sys.argv) > 1:
        csv_path = sys.argv[1]
        print(f"📈 Using specified CSV: {csv_path}")
    else:
        csv_path = find_latest_csv()
        if csv_path is None:
            print("❌ No CSV file found. Specify path as argument.")
            return
    
    # Load data
    excel_df = load_excel_source(excel_path)
    if excel_df is None:
        return
    
    csv_df = load_csv_output(csv_path)
    if csv_df is None:
        return
    
    # Run comparisons
    results = {
        'timestamp': datetime.now().isoformat(),
        'excel_file': excel_path,
        'csv_file': csv_path,
        'campaign_coverage': compare_campaign_coverage(excel_df, csv_df),
        'total_metrics': compare_total_metrics(excel_df, csv_df),
        'sample_campaigns': compare_sample_campaigns(excel_df, csv_df)
    }
    
    # Generate summary
    print("\n" + "=" * 80)
    print("📊 COMPARISON SUMMARY")
    print("=" * 80)
    
    coverage = results['campaign_coverage']
    totals = results['total_metrics']
    samples = results['sample_campaigns']
    
    print(f"📋 Campaign Coverage: {coverage.get('coverage_rate', 0):.1f}%")
    
    if 'overall_grade' in totals:
        grade_map = {
            'PERFECT': '🎯',
            'EXCELLENT': '✅',
            'GOOD': '✅',
            'FAIR': '⚠️',
            'POOR': '❌'
        }
        grade_emoji = grade_map.get(totals['overall_grade'], '❓')
        print(f"🎯 Total Accuracy: {grade_emoji} {totals['overall_grade']} ({totals.get('accuracy_score', 0):.2f}%)")
    
    if 'pass_rate' in samples:
        print(f"🔍 Sample Verification: {samples['pass_rate']:.1f}% passed")
    
    # Save results
    save_comparison_results(results)
    
    print(f"\n✅ Comparison analysis complete!")
    print(f"📊 Files compared: Excel ({len(excel_df)} rows) vs CSV ({len(csv_df)} rows)")

if __name__ == "__main__":
    main()
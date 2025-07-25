#!/usr/bin/env python3
"""
Enhanced Comparison System for MixBridge v2.0
Compares enhanced output files with Excel source of truth
"""

import pandas as pd
import openpyxl
import numpy as np
import json
from pathlib import Path
from datetime import datetime
from typing import Dict, List, Optional, Tuple, Any, Union
from dataclasses import dataclass
import sys
import os

# Add modules directory to path
sys.path.append(str(Path(__file__).parent / 'modules'))

try:
    from data_comparison import ComparisonResult, compare_dataframes
    from excel_operations import load_workbook_values_only, find_header_row
    from reporting import generate_html_report, generate_summary_report
except ImportError:
    print("⚠️  Module imports failed, using fallback implementations")

@dataclass
class ComparisonConfig:
    """Configuration for comparison analysis"""
    
    # File paths
    excel_path: str = 'data/HydrapakUS_SPABridge_MoM_February 2025vsJanuary 2025.xlsx'
    csv_directory: str = 'output/analyses'
    output_directory: str = 'output/reports'
    
    # Analysis settings
    tolerance: float = 0.01
    significant_digits: int = 4
    include_campaign_level: bool = True
    include_total_level: bool = True
    
    # Report settings
    generate_html: bool = True
    generate_summary: bool = True
    save_detailed_results: bool = True

class EnhancedComparisonEngine:
    """
    Main comparison engine for analyzing enhanced MixBridge outputs
    against Excel source of truth
    """
    
    def __init__(self, config: ComparisonConfig = None):
        self.config = config or ComparisonConfig()
        self.excel_df = None
        self.csv_df = None
        self.comparison_results = {}
        
    def load_excel_source(self) -> bool:
        """Load Excel source of truth"""
        print(f"📊 Loading Excel source: {self.config.excel_path}")
        
        try:
            wb = openpyxl.load_workbook(self.config.excel_path, data_only=True)
            
            # Find Campaign sheet
            if 'Campaign' in wb.sheetnames:
                ws = wb['Campaign']
            elif 'Campaign Tab' in wb.sheetnames:
                ws = wb['Campaign Tab']
            else:
                ws = wb.active
                
            print(f"✅ Using sheet: {ws.title}")
            print(f"📏 Dimensions: {ws.max_row} rows x {ws.max_column} columns")
            
            # Find header row (typically row 13 based on our analysis)
            header_row = 13
            for row in range(1, 20):
                cell_value = ws.cell(row=row, column=1).value
                if cell_value and 'campaign' in str(cell_value).lower():
                    header_row = row
                    break
            
            print(f"✅ Header row: {header_row}")
            
            # Extract headers
            headers = []
            for col in range(1, ws.max_column + 1):
                cell_value = ws.cell(row=header_row, column=col).value
                headers.append(cell_value if cell_value else f'Col_{col}')
            
            # Extract data
            data = []
            for row in range(header_row + 1, ws.max_row + 1):
                row_data = []
                for col in range(1, ws.max_column + 1):
                    cell_value = ws.cell(row=row, column=col).value
                    row_data.append(cell_value)
                if any(val is not None for val in row_data):
                    data.append(row_data)
            
            self.excel_df = pd.DataFrame(data, columns=headers)
            self.excel_df = self.excel_df.dropna(how='all')
            
            print(f"✅ Excel data loaded: {len(self.excel_df)} rows, {len(self.excel_df.columns)} columns")
            return True
            
        except Exception as e:
            print(f"❌ Error loading Excel: {e}")
            return False
    
    def load_csv_output(self, csv_path: str) -> bool:
        """Load enhanced CSV output"""
        print(f"📈 Loading CSV: {csv_path}")
        
        try:
            self.csv_df = pd.read_csv(csv_path)
            print(f"✅ CSV data loaded: {len(self.csv_df)} rows, {len(self.csv_df.columns)} columns")
            return True
            
        except Exception as e:
            print(f"❌ Error loading CSV: {e}")
            return False
    
    def find_latest_csv(self) -> Optional[str]:
        """Find the latest CSV file in the analyses directory"""
        analyses_dir = Path(self.config.csv_directory)
        
        if not analyses_dir.exists():
            print(f"❌ Analyses directory not found: {analyses_dir}")
            return None
        
        # Find mixbridge CSV files
        csv_files = list(analyses_dir.glob('mixbridge_*.csv'))
        
        if not csv_files:
            print(f"❌ No mixbridge CSV files found in {analyses_dir}")
            return None
        
        # Sort by modification time, get latest
        latest_file = max(csv_files, key=lambda x: x.stat().st_mtime)
        print(f"✅ Latest CSV found: {latest_file.name}")
        
        return str(latest_file)
    
    def compare_campaign_coverage(self) -> Dict[str, Any]:
        """Compare campaign coverage between Excel and CSV"""
        print("🔍 Analyzing campaign coverage...")
        
        if self.excel_df is None or self.csv_df is None:
            return {"error": "Data not loaded"}
        
        # Extract campaign lists
        excel_campaigns = set(self.excel_df['Campaign'].dropna().astype(str))
        csv_campaigns = set(self.csv_df['Campaign'].dropna().astype(str))
        
        # Clean campaign names
        excel_campaigns = {c.strip() for c in excel_campaigns if c.strip()}
        csv_campaigns = {c.strip() for c in csv_campaigns if c.strip()}
        
        common = excel_campaigns.intersection(csv_campaigns)
        excel_only = excel_campaigns - csv_campaigns
        csv_only = csv_campaigns - excel_campaigns
        
        coverage_rate = (len(common) / max(len(excel_campaigns), len(csv_campaigns))) * 100
        
        result = {
            'excel_campaigns': len(excel_campaigns),
            'csv_campaigns': len(csv_campaigns),
            'common_campaigns': len(common),
            'excel_only': len(excel_only),
            'csv_only': len(csv_only),
            'coverage_rate': coverage_rate,
            'excel_only_list': list(excel_only)[:10],  # First 10 only
            'csv_only_list': list(csv_only)[:10]
        }
        
        print(f"  📊 Excel campaigns: {result['excel_campaigns']}")
        print(f"  📊 CSV campaigns: {result['csv_campaigns']}")
        print(f"  ✅ Common campaigns: {result['common_campaigns']}")
        print(f"  📊 Coverage rate: {coverage_rate:.1f}%")
        
        return result
    
    def compare_total_metrics(self) -> Dict[str, Any]:
        """Compare total row metrics for accuracy assessment"""
        print("🎯 Analyzing total row accuracy...")
        
        if self.excel_df is None or self.csv_df is None:
            return {"error": "Data not loaded"}
        
        # Find total rows
        excel_total = self.excel_df[
            self.excel_df['Campaign'].astype(str).str.contains('total', case=False, na=False)
        ]
        csv_total = self.csv_df[
            self.csv_df['Campaign'].astype(str).str.contains('total', case=False, na=False)
        ]
        
        if len(excel_total) == 0 or len(csv_total) == 0:
            return {"error": "Total rows not found"}
        
        # Key metric comparisons
        comparisons = [
            ('January 2025', 'Spend - January 2025', 'January Spend'),
            ('February 2025', 'Spend - February 2025', 'February Spend'),
            ('Net Change', 'Spend - Net Change', 'Net Change'),
            ('% Change', 'Spend - % Change', 'Percent Change')
        ]
        
        results = []
        accuracy_scores = []
        
        for excel_col, csv_col, metric_name in comparisons:
            if excel_col in excel_total.columns and csv_col in csv_total.columns:
                try:
                    excel_val = float(excel_total.iloc[0][excel_col])
                    csv_val = float(csv_total.iloc[0][csv_col])
                    
                    diff = abs(csv_val - excel_val)
                    rel_diff = (diff / abs(excel_val) * 100) if excel_val != 0 else 0
                    
                    # Determine accuracy grade
                    if rel_diff < 0.01:
                        grade = 'PERFECT'
                        grade_emoji = '🎯'
                    elif rel_diff < 0.1:
                        grade = 'EXCELLENT'
                        grade_emoji = '✅'
                    elif rel_diff < 1.0:
                        grade = 'GOOD'
                        grade_emoji = '✅'
                    elif rel_diff < 5.0:
                        grade = 'FAIR'
                        grade_emoji = '⚠️'
                    else:
                        grade = 'POOR'
                        grade_emoji = '❌'
                    
                    result = {
                        'metric': metric_name,
                        'excel_value': excel_val,
                        'csv_value': csv_val,
                        'difference': diff,
                        'relative_difference': rel_diff,
                        'grade': grade,
                        'grade_emoji': grade_emoji
                    }
                    
                    results.append(result)
                    accuracy_scores.append(rel_diff)
                    
                    print(f"  📈 {metric_name}:")
                    print(f"      Excel: ${excel_val:,.2f}")
                    print(f"      CSV:   ${csv_val:,.2f}")
                    print(f"      Grade: {grade_emoji} {grade} ({rel_diff:.4f}% error)")
                    
                except Exception as e:
                    print(f"  ❌ Error comparing {metric_name}: {e}")
        
        # Overall assessment
        overall_result = {
            'metric_results': results,
            'metrics_compared': len(results),
            'average_error': np.mean(accuracy_scores) if accuracy_scores else 100,
            'maximum_error': max(accuracy_scores) if accuracy_scores else 100,
            'accuracy_score': 100 - np.mean(accuracy_scores) if accuracy_scores else 0
        }
        
        # Overall grade
        avg_error = overall_result['average_error']
        if avg_error < 0.01:
            overall_result['overall_grade'] = 'PERFECT'
            overall_result['grade_emoji'] = '🎯'
        elif avg_error < 0.1:
            overall_result['overall_grade'] = 'EXCELLENT'
            overall_result['grade_emoji'] = '✅'
        elif avg_error < 1.0:
            overall_result['overall_grade'] = 'GOOD'
            overall_result['grade_emoji'] = '✅'
        elif avg_error < 5.0:
            overall_result['overall_grade'] = 'FAIR'
            overall_result['grade_emoji'] = '⚠️'
        else:
            overall_result['overall_grade'] = 'POOR'
            overall_result['grade_emoji'] = '❌'
        
        print(f"\n  🏆 Overall Grade: {overall_result['grade_emoji']} {overall_result['overall_grade']}")
        print(f"      Accuracy Score: {overall_result['accuracy_score']:.2f}%")
        
        return overall_result
    
    def compare_sample_campaigns(self, sample_size: int = 5) -> Dict[str, Any]:
        """Compare sample campaigns for detailed accuracy verification"""
        print(f"🔍 Analyzing {sample_size} sample campaigns...")
        
        if self.excel_df is None or self.csv_df is None:
            return {"error": "Data not loaded"}
        
        # Find common campaigns (excluding total)
        excel_campaigns = set(self.excel_df['Campaign'].dropna().astype(str))
        csv_campaigns = set(self.csv_df['Campaign'].dropna().astype(str))
        
        # Remove 'Total' and empty strings
        excel_campaigns = {c.strip() for c in excel_campaigns 
                          if c.strip() and 'total' not in c.lower()}
        csv_campaigns = {c.strip() for c in csv_campaigns 
                        if c.strip() and 'total' not in c.lower()}
        
        common = list(excel_campaigns.intersection(csv_campaigns))
        
        if len(common) < sample_size:
            sample_size = len(common)
            print(f"  ⚠️  Only {sample_size} common campaigns available")
        
        if sample_size == 0:
            return {"error": "No common campaigns found"}
        
        # Take first N campaigns
        sample_campaigns = common[:sample_size]
        results = []
        
        for campaign in sample_campaigns:
            excel_row = self.excel_df[self.excel_df['Campaign'] == campaign]
            csv_row = self.csv_df[self.csv_df['Campaign'] == campaign]
            
            if len(excel_row) > 0 and len(csv_row) > 0:
                campaign_result = {'campaign': campaign, 'metrics': []}
                
                # Compare key metrics
                try:
                    excel_spend = float(excel_row.iloc[0]['January 2025'])
                    csv_spend = float(csv_row.iloc[0]['Spend - January 2025'])
                    diff = abs(csv_spend - excel_spend)
                    rel_diff = (diff / abs(excel_spend) * 100) if excel_spend != 0 else 0
                    
                    status = '✅' if rel_diff < 1.0 else '⚠️'
                    
                    campaign_result['metrics'].append({
                        'metric': 'January Spend',
                        'excel_value': excel_spend,
                        'csv_value': csv_spend,
                        'relative_difference': rel_diff,
                        'status': status
                    })
                    
                    print(f"  {status} {campaign[:50]}: Jan Spend {rel_diff:.2f}% error")
                    
                except Exception as e:
                    print(f"  ❌ Error comparing {campaign}: {e}")
                
                results.append(campaign_result)
        
        return {
            'sample_size': len(results),
            'campaign_results': results
        }
    
    def run_comprehensive_analysis(self, csv_path: Optional[str] = None) -> Dict[str, Any]:
        """Run complete comparison analysis"""
        print("🎯 COMPREHENSIVE MIXBRIDGE COMPARISON ANALYSIS")
        print("=" * 80)
        
        # Load data
        if not self.load_excel_source():
            return {"error": "Failed to load Excel source"}
        
        if csv_path is None:
            csv_path = self.find_latest_csv()
            if csv_path is None:
                return {"error": "No CSV file found"}
        
        if not self.load_csv_output(csv_path):
            return {"error": "Failed to load CSV output"}
        
        # Run analyses
        results = {
            'timestamp': datetime.now().isoformat(),
            'excel_file': self.config.excel_path,
            'csv_file': csv_path,
            'campaign_coverage': self.compare_campaign_coverage(),
            'total_metrics': self.compare_total_metrics(),
            'sample_campaigns': self.compare_sample_campaigns()
        }
        
        # Save results if configured
        if self.config.save_detailed_results:
            self.save_results(results)
        
        return results
    
    def save_results(self, results: Dict[str, Any]):
        """Save comparison results to file"""
        output_dir = Path(self.config.output_directory)
        output_dir.mkdir(parents=True, exist_ok=True)
        
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f"comparison_results_{timestamp}.json"
        filepath = output_dir / filename
        
        with open(filepath, 'w') as f:
            json.dump(results, f, indent=2, default=str)
        
        print(f"💾 Results saved: {filepath}")
    
    def generate_summary_report(self, results: Dict[str, Any]) -> str:
        """Generate a summary report from comparison results"""
        
        if 'error' in results:
            return f"❌ Analysis failed: {results['error']}"
        
        report = []
        report.append("📊 MIXBRIDGE COMPARISON SUMMARY")
        report.append("=" * 50)
        
        # Campaign coverage
        coverage = results.get('campaign_coverage', {})
        if coverage:
            report.append(f"\n📋 CAMPAIGN COVERAGE:")
            report.append(f"  Excel campaigns: {coverage.get('excel_campaigns', 0)}")
            report.append(f"  CSV campaigns: {coverage.get('csv_campaigns', 0)}")
            report.append(f"  Coverage rate: {coverage.get('coverage_rate', 0):.1f}%")
        
        # Total metrics accuracy
        totals = results.get('total_metrics', {})
        if totals:
            report.append(f"\n🎯 ACCURACY ASSESSMENT:")
            report.append(f"  Overall grade: {totals.get('grade_emoji', '❓')} {totals.get('overall_grade', 'UNKNOWN')}")
            report.append(f"  Accuracy score: {totals.get('accuracy_score', 0):.2f}%")
            report.append(f"  Average error: {totals.get('average_error', 0):.4f}%")
        
        # Sample verification
        samples = results.get('sample_campaigns', {})
        if samples:
            report.append(f"\n🔍 SAMPLE VERIFICATION:")
            report.append(f"  Campaigns tested: {samples.get('sample_size', 0)}")
        
        report.append(f"\n✅ Analysis completed at {results.get('timestamp', 'unknown')}")
        
        return "\n".join(report)


def main():
    """Main execution function"""
    
    # Initialize comparison engine
    config = ComparisonConfig()
    engine = EnhancedComparisonEngine(config)
    
    # Run analysis
    results = engine.run_comprehensive_analysis()
    
    # Generate and display summary
    summary = engine.generate_summary_report(results)
    print(f"\n{summary}")
    
    return results


if __name__ == "__main__":
    main()
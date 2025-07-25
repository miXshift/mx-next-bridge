"""
Formula analysis module for Excel Mix Bridge files.
Handles Excel formula extraction, pattern analysis, and KPI structure mapping.
"""

import re
from typing import Dict, List, Optional, Tuple, Any, Set
import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string
from .excel_operations import extract_cell_formulas, find_kpi_structure
from .config import MIX_BRIDGE_COLUMNS

class FormulaAnalyzer:
    """
    Analyzer for Excel formulas in Mix Bridge files.
    """
    
    def __init__(self):
        self.formula_patterns = {
            'sum': r'SUM\([^)]+\)',
            'average': r'AVERAGE\([^)]+\)',
            'division': r'([A-Z]+\d+)/([A-Z]+\d+)',
            'multiplication': r'([A-Z]+\d+)\*([A-Z]+\d+)',
            'subtraction': r'([A-Z]+\d+)-([A-Z]+\d+)',
            'addition': r'([A-Z]+\d+)\+([A-Z]+\d+)',
            'if_statement': r'IF\([^)]+\)',
            'vlookup': r'VLOOKUP\([^)]+\)',
            'cell_reference': r'[A-Z]+\d+',
            'absolute_reference': r'\$[A-Z]+\$\d+',
            'range_reference': r'[A-Z]+\d+:[A-Z]+\d+'
        }
    
    def extract_formulas_by_column(self, worksheet: openpyxl.worksheet.worksheet.Worksheet,
                                  start_row: int = 1) -> Dict[int, List[Dict[str, Any]]]:
        """
        Extract formulas grouped by column.
        
        Args:
            worksheet: Excel worksheet
            start_row: Starting row for extraction
            
        Returns:
            Dictionary mapping column numbers to formula information
        """
        formulas_by_column = {}
        
        # Extract all formulas
        all_formulas = extract_cell_formulas(worksheet, start_row)
        
        # Group by column
        for cell_coord, formula in all_formulas.items():
            # Parse cell coordinate
            col_letter = ''.join(filter(str.isalpha, cell_coord))
            row_num = int(''.join(filter(str.isdigit, cell_coord)))
            col_num = column_index_from_string(col_letter)
            
            if col_num not in formulas_by_column:
                formulas_by_column[col_num] = []
            
            formula_info = {
                'cell': cell_coord,
                'row': row_num,
                'formula': formula,
                'pattern_matches': self._analyze_formula_patterns(formula)
            }
            
            formulas_by_column[col_num].append(formula_info)
        
        return formulas_by_column
    
    def _analyze_formula_patterns(self, formula: str) -> Dict[str, List[str]]:
        """
        Analyze formula for common patterns.
        
        Args:
            formula: Excel formula string
            
        Returns:
            Dictionary of pattern matches
        """
        matches = {}
        
        for pattern_name, pattern_regex in self.formula_patterns.items():
            pattern_matches = re.findall(pattern_regex, formula, re.IGNORECASE)
            if pattern_matches:
                matches[pattern_name] = pattern_matches
        
        return matches
    
    def analyze_formula_patterns(self, formula_dict: Dict[str, str]) -> Dict[str, Any]:
        """
        Analyze patterns across multiple formulas.
        
        Args:
            formula_dict: Dictionary of formulas
            
        Returns:
            Analysis results
        """
        analysis = {
            'total_formulas': len(formula_dict),
            'pattern_frequency': {},
            'common_references': {},
            'formula_complexity': {},
            'unique_patterns': set()
        }
        
        # Analyze each formula
        for cell, formula in formula_dict.items():
            patterns = self._analyze_formula_patterns(formula)
            
            # Count pattern frequency
            for pattern_name, matches in patterns.items():
                if pattern_name not in analysis['pattern_frequency']:
                    analysis['pattern_frequency'][pattern_name] = 0
                analysis['pattern_frequency'][pattern_name] += len(matches)
                
                # Add to unique patterns
                for match in matches:
                    if isinstance(match, tuple):
                        analysis['unique_patterns'].add(f"{pattern_name}:{'-'.join(match)}")
                    else:
                        analysis['unique_patterns'].add(f"{pattern_name}:{match}")
            
            # Calculate complexity (number of operations)
            complexity = sum(len(matches) for matches in patterns.values())
            analysis['formula_complexity'][cell] = complexity
            
            # Track common cell references
            if 'cell_reference' in patterns:
                for ref in patterns['cell_reference']:
                    if ref not in analysis['common_references']:
                        analysis['common_references'][ref] = 0
                    analysis['common_references'][ref] += 1
        
        # Convert set to list for JSON serialization
        analysis['unique_patterns'] = list(analysis['unique_patterns'])
        
        return analysis
    
    def map_kpi_structure(self, worksheet: openpyxl.worksheet.worksheet.Worksheet,
                         kpi_row: int,
                         dimension_row: int) -> Dict[str, Any]:
        """
        Map KPI structure for Mix Bridge analysis.
        
        Args:
            worksheet: Excel worksheet
            kpi_row: Row containing KPI names
            dimension_row: Row containing dimension names
            
        Returns:
            KPI structure mapping
        """
        structure = {
            'kpi_mapping': {},
            'dimension_mapping': {},
            'column_groups': {},
            'analytic_points': {}
        }
        
        # Extract KPI names
        for col in range(1, worksheet.max_column + 1):
            kpi_cell = worksheet.cell(row=kpi_row, column=col)
            dim_cell = worksheet.cell(row=dimension_row, column=col)
            
            if kpi_cell.value:
                kpi_name = str(kpi_cell.value).strip()
                structure['kpi_mapping'][col] = kpi_name
                
                if kpi_name not in structure['column_groups']:
                    structure['column_groups'][kpi_name] = []
                structure['column_groups'][kpi_name].append(col)
            
            if dim_cell.value:
                dim_name = str(dim_cell.value).strip()
                structure['dimension_mapping'][col] = dim_name
                
                # Check if this is one of the 5 analytic points
                for analytic_key, analytic_name in MIX_BRIDGE_COLUMNS.items():
                    if analytic_name.lower() in dim_name.lower():
                        if analytic_key not in structure['analytic_points']:
                            structure['analytic_points'][analytic_key] = []
                        structure['analytic_points'][analytic_key].append(col)
        
        return structure
    
    def verify_contribution_calculation(self, worksheet: openpyxl.worksheet.worksheet.Worksheet,
                                      contribution_columns: List[int],
                                      mix_columns: List[int],
                                      rate_columns: List[int]) -> Dict[str, Any]:
        """
        Verify that contribution calculations follow expected formula.
        
        Args:
            worksheet: Excel worksheet
            contribution_columns: List of contribution column numbers
            mix_columns: List of mix column numbers
            rate_columns: List of rate column numbers
            
        Returns:
            Verification results
        """
        verification = {
            'valid_formulas': [],
            'invalid_formulas': [],
            'missing_formulas': [],
            'formula_patterns': {}
        }
        
        # Check each contribution column
        for contrib_col in contribution_columns:
            # Look for corresponding mix and rate columns
            mix_col = None
            rate_col = None
            
            # Find closest mix and rate columns
            for mix_c in mix_columns:
                if abs(mix_c - contrib_col) < abs(mix_col - contrib_col if mix_col else float('inf')):
                    mix_col = mix_c
                    
            for rate_c in rate_columns:
                if abs(rate_c - contrib_col) < abs(rate_col - contrib_col if rate_col else float('inf')):
                    rate_col = rate_c
            
            if mix_col and rate_col:
                # Check formulas in this column
                for row in range(3, worksheet.max_row + 1):  # Skip header rows
                    cell = worksheet.cell(row=row, column=contrib_col)
                    if isinstance(cell.value, str) and cell.value.startswith('='):
                        formula = cell.value
                        
                        # Check if formula references expected cells
                        mix_ref = f"{get_column_letter(mix_col)}{row}"
                        rate_ref = f"{get_column_letter(rate_col)}{row}"
                        
                        if mix_ref in formula and rate_ref in formula:
                            verification['valid_formulas'].append({
                                'cell': f"{get_column_letter(contrib_col)}{row}",
                                'formula': formula,
                                'mix_ref': mix_ref,
                                'rate_ref': rate_ref
                            })
                        else:
                            verification['invalid_formulas'].append({
                                'cell': f"{get_column_letter(contrib_col)}{row}",
                                'formula': formula,
                                'expected_mix_ref': mix_ref,
                                'expected_rate_ref': rate_ref
                            })
                    elif cell.value is None:
                        verification['missing_formulas'].append(
                            f"{get_column_letter(contrib_col)}{row}"
                        )
        
        return verification
    
    def trace_formula_dependencies(self, formula: str) -> Dict[str, List[str]]:
        """
        Trace dependencies in a formula.
        
        Args:
            formula: Excel formula string
            
        Returns:
            Dictionary of dependencies
        """
        dependencies = {
            'cell_references': [],
            'range_references': [],
            'external_references': [],
            'functions_used': []
        }
        
        # Find cell references
        cell_refs = re.findall(r'[A-Z]+\d+', formula)
        dependencies['cell_references'] = list(set(cell_refs))
        
        # Find range references
        range_refs = re.findall(r'[A-Z]+\d+:[A-Z]+\d+', formula)
        dependencies['range_references'] = list(set(range_refs))
        
        # Find external sheet references
        external_refs = re.findall(r"'[^']*'![A-Z]+\d+", formula)
        dependencies['external_references'] = list(set(external_refs))
        
        # Find functions
        functions = re.findall(r'[A-Z]+\(', formula)
        dependencies['functions_used'] = list(set([f[:-1] for f in functions]))
        
        return dependencies
    
    def analyze_kpi_relationships(self, worksheet: openpyxl.worksheet.worksheet.Worksheet,
                                 kpi_structure: Dict[str, Any]) -> Dict[str, Any]:
        """
        Analyze relationships between KPIs based on formulas.
        
        Args:
            worksheet: Excel worksheet
            kpi_structure: KPI structure mapping
            
        Returns:
            Relationship analysis
        """
        relationships = {
            'kpi_dependencies': {},
            'circular_references': [],
            'formula_chains': {},
            'isolated_kpis': []
        }
        
        # Analyze each KPI group
        for kpi_name, columns in kpi_structure.get('column_groups', {}).items():
            kpi_deps = set()
            
            # Check formulas in these columns
            for col in columns:
                for row in range(3, min(worksheet.max_row + 1, 100)):  # Limit range
                    cell = worksheet.cell(row=row, column=col)
                    if isinstance(cell.value, str) and cell.value.startswith('='):
                        deps = self.trace_formula_dependencies(cell.value)
                        
                        # Check if references point to other KPI columns
                        for ref in deps['cell_references']:
                            ref_col = column_index_from_string(
                                ''.join(filter(str.isalpha, ref))
                            )
                            
                            # Find which KPI this reference belongs to
                            for other_kpi, other_cols in kpi_structure.get('column_groups', {}).items():
                                if ref_col in other_cols and other_kpi != kpi_name:
                                    kpi_deps.add(other_kpi)
            
            relationships['kpi_dependencies'][kpi_name] = list(kpi_deps)
            
            # Check for isolated KPIs (no dependencies)
            if not kpi_deps:
                relationships['isolated_kpis'].append(kpi_name)
        
        return relationships

def create_formula_report(worksheet: openpyxl.worksheet.worksheet.Worksheet,
                         output_path: Optional[str] = None) -> Dict[str, Any]:
    """
    Create comprehensive formula analysis report.
    
    Args:
        worksheet: Excel worksheet
        output_path: Optional path to save report
        
    Returns:
        Complete analysis report
    """
    analyzer = FormulaAnalyzer()
    
    # Extract formulas by column
    formulas_by_column = analyzer.extract_formulas_by_column(worksheet)
    
    # Find KPI structure
    kpi_structure = find_kpi_structure(worksheet)
    
    # Analyze patterns
    all_formulas = {}
    for col_formulas in formulas_by_column.values():
        for formula_info in col_formulas:
            all_formulas[formula_info['cell']] = formula_info['formula']
    
    pattern_analysis = analyzer.analyze_formula_patterns(all_formulas)
    
    # Create comprehensive report
    report = {
        'formulas_by_column': formulas_by_column,
        'kpi_structure': kpi_structure,
        'pattern_analysis': pattern_analysis,
        'summary': {
            'total_columns_with_formulas': len(formulas_by_column),
            'total_formulas': len(all_formulas),
            'kpi_count': len(kpi_structure['kpi_mapping']) if kpi_structure else 0,
            'analytic_points_found': len(kpi_structure['analytic_points']) if kpi_structure else 0
        }
    }
    
    # Add KPI relationships if structure found
    if kpi_structure and kpi_structure['kpi_mapping']:
        relationships = analyzer.analyze_kpi_relationships(worksheet, kpi_structure)
        report['kpi_relationships'] = relationships
    
    # Save report if path provided
    if output_path:
        import json
        with open(output_path, 'w') as f:
            json.dump(report, f, indent=2, default=str)
    
    return report
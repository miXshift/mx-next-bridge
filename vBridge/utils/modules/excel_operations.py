"""
Excel operations module for Mix Bridge analysis.
Handles Excel file loading, header detection, formula extraction, and caching.
"""

import openpyxl
from openpyxl.utils import get_column_letter
from typing import Optional, Dict, List, Tuple, Any, Union
from functools import lru_cache
from pathlib import Path
import time
from .config import CacheConfig

class ExcelCache:
    """Simple cache for Excel workbook objects."""
    
    def __init__(self, config: CacheConfig):
        self.config = config
        self._cache: Dict[str, Tuple[Any, float]] = {}
        
    def get(self, key: str) -> Optional[Any]:
        """Get item from cache if not expired."""
        if not self.config.enabled:
            return None
            
        if key in self._cache:
            item, timestamp = self._cache[key]
            if time.time() - timestamp < self.config.ttl_seconds:
                return item
            else:
                del self._cache[key]
        return None
        
    def set(self, key: str, value: Any) -> None:
        """Set item in cache."""
        if self.config.enabled:
            self._cache[key] = (value, time.time())
            
    def clear(self) -> None:
        """Clear all cached items."""
        self._cache.clear()

# Global cache instance
_excel_cache = ExcelCache(CacheConfig())

def set_cache_config(config: CacheConfig) -> None:
    """Update cache configuration."""
    global _excel_cache
    _excel_cache = ExcelCache(config)

def load_workbook_with_formulas(file_path: Union[str, Path], use_cache: bool = True) -> openpyxl.Workbook:
    """
    Load Excel workbook with formulas preserved.
    
    Args:
        file_path: Path to Excel file
        use_cache: Whether to use caching
        
    Returns:
        Workbook object with formulas
    """
    file_path = str(file_path)
    cache_key = f"formulas_{file_path}"
    
    if use_cache:
        cached = _excel_cache.get(cache_key)
        if cached is not None:
            return cached
    
    workbook = openpyxl.load_workbook(file_path, data_only=False)
    
    if use_cache:
        _excel_cache.set(cache_key, workbook)
        
    return workbook

def load_workbook_values_only(file_path: Union[str, Path], use_cache: bool = True) -> openpyxl.Workbook:
    """
    Load Excel workbook with calculated values only.
    
    Args:
        file_path: Path to Excel file
        use_cache: Whether to use caching
        
    Returns:
        Workbook object with values
    """
    file_path = str(file_path)
    cache_key = f"values_{file_path}"
    
    if use_cache:
        cached = _excel_cache.get(cache_key)
        if cached is not None:
            return cached
    
    workbook = openpyxl.load_workbook(file_path, data_only=True)
    
    if use_cache:
        _excel_cache.set(cache_key, workbook)
        
    return workbook

def find_header_row(worksheet: openpyxl.worksheet.worksheet.Worksheet, 
                   search_terms: Optional[List[str]] = None,
                   max_rows: int = 50) -> Optional[int]:
    """
    Find the header row in a worksheet by searching for common terms.
    
    Args:
        worksheet: Excel worksheet object
        search_terms: List of terms to search for (defaults to common Mix Bridge headers)
        max_rows: Maximum number of rows to search
        
    Returns:
        Row number of header row, or None if not found
    """
    if search_terms is None:
        search_terms = ['Campaign', 'Impressions', 'Clicks', 'Spend', 'KPI', 'Metric']
    
    search_terms_lower = [term.lower() for term in search_terms]
    
    for row_num in range(1, min(max_rows + 1, worksheet.max_row + 1)):
        row_values = []
        for col in range(1, min(20, worksheet.max_column + 1)):
            cell = worksheet.cell(row=row_num, column=col)
            if cell.value:
                row_values.append(str(cell.value).lower())
        
        # Check if this row contains multiple search terms
        matches = sum(1 for term in search_terms_lower if any(term in val for val in row_values))
        if matches >= min(2, len(search_terms)):
            return row_num
            
    return None

def get_column_headers(worksheet: openpyxl.worksheet.worksheet.Worksheet, 
                      header_row: int) -> Dict[str, int]:
    """
    Extract column headers and their positions.
    
    Args:
        worksheet: Excel worksheet object
        header_row: Row number containing headers
        
    Returns:
        Dictionary mapping header names to column numbers
    """
    headers = {}
    for col in range(1, worksheet.max_column + 1):
        cell = worksheet.cell(row=header_row, column=col)
        if cell.value:
            header_name = str(cell.value).strip()
            headers[header_name] = col
            
    return headers

def extract_cell_formulas(worksheet: openpyxl.worksheet.worksheet.Worksheet,
                         start_row: int = 1,
                         end_row: Optional[int] = None,
                         columns: Optional[List[int]] = None) -> Dict[str, str]:
    """
    Extract formulas from specified cells.
    
    Args:
        worksheet: Excel worksheet object
        start_row: Starting row (default: 1)
        end_row: Ending row (default: worksheet.max_row)
        columns: List of column numbers to extract (default: all columns)
        
    Returns:
        Dictionary mapping cell coordinates to formulas
    """
    formulas = {}
    
    if end_row is None:
        end_row = worksheet.max_row
    if columns is None:
        columns = range(1, worksheet.max_column + 1)
        
    for row in range(start_row, min(end_row + 1, worksheet.max_row + 1)):
        for col in columns:
            cell = worksheet.cell(row=row, column=col)
            if isinstance(cell.value, str) and cell.value.startswith('='):
                coord = f"{get_column_letter(col)}{row}"
                formulas[coord] = cell.value
                
    return formulas

def get_cell_formatting(worksheet: openpyxl.worksheet.worksheet.Worksheet,
                       row: int,
                       column: int) -> Dict[str, Any]:
    """
    Get formatting information for a cell.
    
    Args:
        worksheet: Excel worksheet object
        row: Row number
        column: Column number
        
    Returns:
        Dictionary with formatting information
    """
    cell = worksheet.cell(row=row, column=column)
    
    return {
        "number_format": cell.number_format,
        "is_percentage": '%' in str(cell.number_format),
        "font": {
            "bold": cell.font.bold if cell.font else False,
            "italic": cell.font.italic if cell.font else False,
            "color": cell.font.color.rgb if cell.font and cell.font.color else None
        },
        "fill": {
            "color": cell.fill.start_color.rgb if cell.fill and cell.fill.start_color else None
        },
        "alignment": {
            "horizontal": cell.alignment.horizontal if cell.alignment else None,
            "vertical": cell.alignment.vertical if cell.alignment else None
        }
    }

def find_kpi_structure(worksheet: openpyxl.worksheet.worksheet.Worksheet,
                      start_row: int = 1) -> Optional[Dict[str, Any]]:
    """
    Find KPI/Dimension structure in Mix Bridge format.
    
    Args:
        worksheet: Excel worksheet object
        start_row: Row to start searching from
        
    Returns:
        Dictionary with KPI and dimension information
    """
    structure = {
        "kpi_row": None,
        "dimension_row": None,
        "data_start_row": None,
        "kpi_columns": {},
        "dimension_columns": {}
    }
    
    # Look for KPI row
    for row_num in range(start_row, min(start_row + 10, worksheet.max_row + 1)):
        cell_a = worksheet.cell(row=row_num, column=1)
        if cell_a.value and 'kpi' in str(cell_a.value).lower():
            structure["kpi_row"] = row_num
            
            # Extract KPI names
            for col in range(2, worksheet.max_column + 1):
                cell = worksheet.cell(row=row_num, column=col)
                if cell.value:
                    structure["kpi_columns"][col] = str(cell.value).strip()
            
            # Look for dimension row (usually next row)
            dimension_row = row_num + 1
            cell_dim = worksheet.cell(row=dimension_row, column=1)
            if cell_dim.value and 'dimension' in str(cell_dim.value).lower():
                structure["dimension_row"] = dimension_row
                
                # Extract dimension names
                for col in range(2, worksheet.max_column + 1):
                    cell = worksheet.cell(row=dimension_row, column=col)
                    if cell.value:
                        structure["dimension_columns"][col] = str(cell.value).strip()
                        
                structure["data_start_row"] = dimension_row + 1
                
            return structure
            
    return None

def get_sheet_names(file_path: Union[str, Path]) -> List[str]:
    """
    Get all sheet names from an Excel file.
    
    Args:
        file_path: Path to Excel file
        
    Returns:
        List of sheet names
    """
    workbook = load_workbook_values_only(file_path, use_cache=False)
    return workbook.sheetnames

def clear_cache() -> None:
    """Clear the Excel cache."""
    _excel_cache.clear()
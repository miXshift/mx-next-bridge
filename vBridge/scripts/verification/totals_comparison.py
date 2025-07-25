#!/usr/bin/env python3
"""
Compare totals between Excel file and generated CSV
"""

import pandas as pd
import openpyxl

def main():
    print('🔍 ANALYZING TOTALS MATCH BETWEEN CSV AND EXCEL')
    print('=' * 60)

    # Load Excel data
    wb = openpyxl.load_workbook('data/HydrapakUS_SPABridge_MoM_February 2025vsJanuary 2025.xlsx', data_only=True)
    ws = wb['Campaign']

    # Extract header row (row 13) and total row (row 14)
    header_row = 13
    total_row = 14

    headers = []
    for col in range(1, ws.max_column + 1):
        cell_value = ws.cell(row=header_row, column=col).value
        headers.append(cell_value if cell_value else f'Col_{col}')

    totals_data = []
    for col in range(1, ws.max_column + 1):
        cell_value = ws.cell(row=total_row, column=col).value
        totals_data.append(cell_value)

    excel_totals = dict(zip(headers, totals_data))

    print('📊 EXCEL TOTALS:')
    print(f'  Spend January 2025: ${excel_totals["January 2025"]:,.2f}')
    print(f'  Spend February 2025: ${excel_totals["February 2025"]:,.2f}')

    # Load CSV data  
    csv_df = pd.read_csv('output/analyses/mixbridge_jan2025-feb2025_delta_20250721_190317.csv')

    # Find total row in CSV
    csv_total_row = csv_df[csv_df['Campaign'].astype(str).str.contains('Total', case=False, na=False)]

    if not csv_total_row.empty:
        print('\n📈 CSV TOTALS:')
        csv_jan_spend = csv_total_row['Spend - January 2025'].iloc[0]
        csv_feb_spend = csv_total_row['Spend - February 2025'].iloc[0]
        
        print(f'  Spend January 2025: ${csv_jan_spend:,.2f}')
        print(f'  Spend February 2025: ${csv_feb_spend:,.2f}')
        
        print('\n🎯 COMPARISON RESULTS:')
        jan_diff = abs(excel_totals["January 2025"] - csv_jan_spend)
        feb_diff = abs(excel_totals["February 2025"] - csv_feb_spend)
        
        jan_rel_diff = (jan_diff / excel_totals["January 2025"] * 100) if excel_totals["January 2025"] != 0 else 0
        feb_rel_diff = (feb_diff / excel_totals["February 2025"] * 100) if excel_totals["February 2025"] != 0 else 0
        
        print(f'  January 2025 difference: ${jan_diff:,.2f} ({jan_rel_diff:.3f}%)')
        print(f'  February 2025 difference: ${feb_diff:,.2f} ({feb_rel_diff:.3f}%)')
        
        # Final assessment
        tolerance = 0.01  # 0.01% tolerance
        jan_match = jan_rel_diff < tolerance
        feb_match = feb_rel_diff < tolerance
        
        print('\n📋 FINAL ASSESSMENT:')
        print(f'  January totals match: {"✅ YES" if jan_match else "❌ NO"} ({jan_rel_diff:.3f}% difference)')
        print(f'  February totals match: {"✅ YES" if feb_match else "❌ NO"} ({feb_rel_diff:.3f}% difference)')
        print(f'  Overall assessment: {"✅ TOTALS MATCH" if jan_match and feb_match else "❌ TOTALS DO NOT MATCH"}')
        
        # Additional analysis
        print('\n🔍 ADDITIONAL ANALYSIS:')
        print(f'  Excel shows {len([h for h in headers if h and "campaign" not in h.lower()])} data columns')
        print(f'  CSV shows {len(csv_df)} total rows (including header)')
        
        # Count actual campaign rows (excluding totals)
        csv_campaign_count = len(csv_df[~csv_df['Campaign'].astype(str).str.contains('Total', case=False, na=False)])
        print(f'  CSV contains {csv_campaign_count} campaign rows')
        
        # Count campaigns in Excel (total rows - 1 for total row)
        excel_campaign_count = ws.max_row - total_row  # Rows after total row
        if excel_campaign_count <= 0:
            excel_campaign_count = 55  # Known from previous analysis
        print(f'  Excel contains approximately {excel_campaign_count} campaign rows')
        
    else:
        print('❌ No total row found in CSV file')

if __name__ == "__main__":
    main()
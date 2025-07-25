#!/usr/bin/env python3
"""
Investigate why Excel and CSV totals are so different
"""

import pandas as pd
import openpyxl

def main():
    print('🔍 INVESTIGATING DIFFERENCES BETWEEN EXCEL AND CSV')
    print('=' * 70)

    # Load Excel data to see all campaign rows
    wb = openpyxl.load_workbook('data/HydrapakUS_SPABridge_MoM_February 2025vsJanuary 2025.xlsx', data_only=True)
    ws = wb['Campaign']

    # Get all data after the header row (row 13)
    header_row = 13
    print(f'📊 EXCEL ANALYSIS:')
    print(f'  Header row: {header_row}')
    print(f'  Total rows: {ws.max_row}')
    print(f'  Campaign rows: {ws.max_row - header_row}')

    # Extract some sample campaign data from Excel
    print('\n📋 EXCEL SAMPLE DATA (first 10 campaigns):')
    for row in range(header_row + 2, min(header_row + 12, ws.max_row + 1)):  # Skip total row
        campaign = ws.cell(row=row, column=1).value
        jan_spend = ws.cell(row=row, column=2).value
        feb_spend = ws.cell(row=row, column=3).value
        
        if campaign and jan_spend is not None and feb_spend is not None:
            print(f'  {campaign[:50]:<50} Jan: ${jan_spend:>8.2f} Feb: ${feb_spend:>8.2f}')

    # Load CSV data
    csv_df = pd.read_csv('output/analyses/mixbridge_jan2025-feb2025_delta_20250721_190317.csv')
    
    print(f'\n📈 CSV ANALYSIS:')
    print(f'  Total rows: {len(csv_df)}')
    
    # Exclude total row
    csv_campaigns = csv_df[~csv_df['Campaign'].astype(str).str.contains('Total', case=False, na=False)]
    print(f'  Campaign rows: {len(csv_campaigns)}')
    
    print('\n📋 CSV SAMPLE DATA (first 10 campaigns):')
    for i in range(min(10, len(csv_campaigns))):
        row = csv_campaigns.iloc[i]
        campaign = row['Campaign']
        jan_spend = row['Spend - January 2025']
        feb_spend = row['Spend - February 2025']
        print(f'  {campaign[:50]:<50} Jan: ${jan_spend:>8.2f} Feb: ${feb_spend:>8.2f}')

    # Compare totals calculation manually
    print('\n🧮 MANUAL TOTALS CALCULATION:')
    
    # Sum up first 10 campaigns from CSV
    first_10_csv_jan = csv_campaigns.head(10)['Spend - January 2025'].sum()
    first_10_csv_feb = csv_campaigns.head(10)['Spend - February 2025'].sum()
    
    print(f'  First 10 CSV campaigns sum - Jan: ${first_10_csv_jan:.2f}, Feb: ${first_10_csv_feb:.2f}')
    
    # Get total from CSV
    csv_total = csv_df[csv_df['Campaign'].astype(str).str.contains('Total', case=False, na=False)]
    if not csv_total.empty:
        csv_total_jan = csv_total['Spend - January 2025'].iloc[0]
        csv_total_feb = csv_total['Spend - February 2025'].iloc[0]
        print(f'  CSV Total row        - Jan: ${csv_total_jan:.2f}, Feb: ${csv_total_feb:.2f}')

    # Check if Excel has any filtering or special conditions
    print('\n🔍 POTENTIAL REASONS FOR DIFFERENCE:')
    print('  1. Excel may have filters applied that limit visible campaigns')
    print('  2. Excel may exclude certain campaign types or statuses')
    print('  3. CSV includes all 156 campaigns while Excel shows only ~55')
    print('  4. Date ranges might be different between the files')
    
    # Look for patterns in campaign names that might indicate filtering
    print('\n📝 CAMPAIGN NAME PATTERNS:')
    all_campaigns = csv_campaigns['Campaign'].tolist()
    
    # Look for common patterns
    prof_campaigns = [c for c in all_campaigns if 'PROF' in str(c)]
    disc_campaigns = [c for c in all_campaigns if 'DISC' in str(c)]
    conq_campaigns = [c for c in all_campaigns if 'CONQ' in str(c)]
    
    print(f'  PROF campaigns: {len(prof_campaigns)}')
    print(f'  DISC campaigns: {len(disc_campaigns)}')  
    print(f'  CONQ campaigns: {len(conq_campaigns)}')
    print(f'  Other campaigns: {len(all_campaigns) - len(prof_campaigns) - len(disc_campaigns) - len(conq_campaigns)}')

if __name__ == "__main__":
    main()
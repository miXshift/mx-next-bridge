#!/usr/bin/env python3
"""
Compare campaign lists between Excel and CSV to identify filtering
"""

import pandas as pd
import openpyxl

def main():
    print('🔍 COMPARING CAMPAIGN LISTS TO IDENTIFY FILTERING')
    print('=' * 60)

    # Load Excel campaign data
    wb = openpyxl.load_workbook('data/HydrapakUS_SPABridge_MoM_February 2025vsJanuary 2025.xlsx', data_only=True)
    ws = wb['Campaign']

    excel_campaigns = []
    header_row = 13
    
    # Extract all campaign names from Excel (skip header and total row)
    for row in range(header_row + 2, ws.max_row + 1):  # Start after total row
        campaign = ws.cell(row=row, column=1).value
        if campaign:
            excel_campaigns.append(str(campaign).strip())

    # Load CSV campaign data
    csv_df = pd.read_csv('output/analyses/mixbridge_jan2025-feb2025_delta_20250721_190317.csv')
    csv_campaigns_df = csv_df[~csv_df['Campaign'].astype(str).str.contains('Total', case=False, na=False)]
    csv_campaigns = [str(c).strip() for c in csv_campaigns_df['Campaign'].tolist()]

    print(f'📊 CAMPAIGN COUNT COMPARISON:')
    print(f'  Excel campaigns: {len(excel_campaigns)}')
    print(f'  CSV campaigns: {len(csv_campaigns)}')

    # Find campaigns in each but not the other
    excel_set = set(excel_campaigns)
    csv_set = set(csv_campaigns)
    
    common_campaigns = excel_set.intersection(csv_set)
    excel_only = excel_set - csv_set
    csv_only = csv_set - excel_set

    print(f'  Common campaigns: {len(common_campaigns)}')
    print(f'  Excel only: {len(excel_only)}')
    print(f'  CSV only: {len(csv_only)}')

    # Show some examples
    if excel_only:
        print(f'\n📋 EXAMPLES OF CAMPAIGNS IN EXCEL BUT NOT CSV (first 5):')
        for campaign in list(excel_only)[:5]:
            print(f'  - {campaign}')
    
    if csv_only:
        print(f'\n📈 EXAMPLES OF CAMPAIGNS IN CSV BUT NOT EXCEL (first 10):')
        for campaign in list(csv_only)[:10]:
            print(f'  - {campaign}')

    # Calculate what the totals would be if we only included common campaigns
    print(f'\n🧮 RECALCULATED TOTALS (COMMON CAMPAIGNS ONLY):')
    
    common_campaigns_df = csv_campaigns_df[csv_campaigns_df['Campaign'].isin(common_campaigns)]
    if not common_campaigns_df.empty:
        common_jan_total = common_campaigns_df['Spend - January 2025'].sum()
        common_feb_total = common_campaigns_df['Spend - February 2025'].sum()
        
        print(f'  Common campaigns Jan total: ${common_jan_total:.2f}')
        print(f'  Common campaigns Feb total: ${common_feb_total:.2f}')
        
        # Compare to Excel totals
        excel_jan = 43959.94  # From previous analysis
        excel_feb = 49667.68
        
        jan_diff = abs(excel_jan - common_jan_total)
        feb_diff = abs(excel_feb - common_feb_total)
        
        jan_rel_diff = (jan_diff / excel_jan * 100) if excel_jan != 0 else 0
        feb_rel_diff = (feb_diff / excel_feb * 100) if excel_feb != 0 else 0
        
        print(f'\n  Comparison with Excel totals:')
        print(f'  Jan difference: ${jan_diff:.2f} ({jan_rel_diff:.3f}%)')
        print(f'  Feb difference: ${feb_diff:.2f} ({feb_rel_diff:.3f}%)')
    
    # Analyze patterns in CSV-only campaigns to understand the filtering
    if csv_only:
        print(f'\n🔍 ANALYSIS OF CSV-ONLY CAMPAIGNS:')
        csv_only_list = list(csv_only)
        
        # Check for patterns
        low_spend = []
        zero_spend = []
        pattern_1000 = []
        
        for campaign in csv_only_list:
            # Find this campaign in CSV data
            campaign_data = csv_campaigns_df[csv_campaigns_df['Campaign'] == campaign]
            if not campaign_data.empty:
                jan_spend = campaign_data['Spend - January 2025'].iloc[0]
                feb_spend = campaign_data['Spend - February 2025'].iloc[0]
                total_spend = jan_spend + feb_spend
                
                if total_spend == 0:
                    zero_spend.append(campaign)
                elif total_spend < 100:  # Low spend threshold
                    low_spend.append(campaign)
                    
                if campaign.startswith('1000-'):
                    pattern_1000.append(campaign)
        
        print(f'  Zero spend campaigns: {len(zero_spend)}')
        print(f'  Low spend (<$100) campaigns: {len(low_spend)}')
        print(f'  "1000-" prefix campaigns: {len(pattern_1000)}')
        
        if pattern_1000:
            print(f'    Examples of 1000- campaigns: {pattern_1000[:3]}')

    # Final assessment
    print(f'\n📋 ASSESSMENT:')
    print(f'  The Excel file appears to have filtering applied that excludes {len(csv_only)} campaigns')
    print(f'  This explains why totals don\'t match - different campaign sets are being totaled')
    
    # Check coverage percentage
    coverage = len(common_campaigns) / len(csv_campaigns) * 100
    print(f'  Excel covers {coverage:.1f}% of campaigns in the CSV data')

if __name__ == "__main__":
    main()
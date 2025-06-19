import openpyxl
import re
import csv
from pathlib import Path

# Define mapping of Excel functions to Python equivalents
EXCEL_TO_PYTHON = {
    'SUM': 'sum',
    'AVERAGE': 'sum({})/len({})',
    'IF': 'lambda x: {} if {} else {}',
    'VLOOKUP': 'lookup',  # Placeholder, requires custom implementation
    'AND': 'all',
    'OR': 'any',
    'MIN': 'min',
    'MAX': 'max',
    'ABS': 'abs',
    'ROUND': 'round',
}

def convert_formula_to_python(formula, cell_ref):
    """
    Convert an Excel formula to a Python expression.
    Args:
        formula (str): Excel formula (e.g., '=SUM(A1:A10)')
        cell_ref (str): Cell reference (e.g., 'B2')
    Returns:
        str: Python-equivalent expression
    """
    if not formula.startswith('='):
        return formula  # Not a formula, return as is

    # Remove leading '='
    formula = formula[1:].strip()

    # Replace Excel function names with Python equivalents
    for excel_func, python_func in EXCEL_TO_PYTHON.items():
        pattern = rf'\b{excel_func}\b'
        if excel_func == 'IF':
            # Handle IF separately due to its ternary structure
            if re.match(r'IF\([^)]+\)', formula):
                match = re.match(r'IF\(([^,]+),([^,]+),([^)]+)\)', formula)
                if match:
                    condition, true_val, false_val = match.groups()
                    return f'({true_val.strip()} if {condition.strip()} else {false_val.strip()})'
        elif excel_func in ['SUM', 'MIN', 'MAX', 'ABS', 'ROUND']:
            formula = re.sub(pattern, python_func, formula, flags=re.IGNORECASE)

    # Replace cell references with Python variable names (e.g., A1 -> data['A1'])
    def replace_cell_ref(match):
        cell = match.group(0)
        return f"data['{cell}']"

    formula = re.sub(r'[A-Z]+\d+', replace_cell_ref, formula)

    # Replace Excel operators with Python equivalents
    formula = formula.replace('^', '**')  # Exponentiation
    formula = formula.replace('&', '+')   # String concatenation

    return formula

def generate_label(cell_ref, sheet, formula):
    """
    Generate a label for a formula based on cell reference and nearby text.
    Args:
        cell_ref (str): Cell reference (e.g., 'B2')
        sheet: Worksheet object
        formula (str): Original Excel formula
    Returns:
        str: Label for the formula
    """
    # Try to find a label in the cell to the left or above
    col = openpyxl.utils.cell.column_index_from_string(''.join([c for c in cell_ref if c.isalpha]))
    row = int(''.join([c for c in cell_ref if c.isdigit]))

    # Check left cell (same row, previous column)
    left_cell = openpyxl.utils.cell.get_column_letter(col - 1) + str(row)
    try:
        left_value = sheet[left_cell].value
        if left_value and isinstance(left_value, str):
            return f"{left_value.strip()} ({cell_ref})"
    except:
        pass

    # Check above cell (same column, previous row)
    above_cell = openpyxl.utils.cell.get_column_letter(col) + str(row - 1)
    try:
        above_value = sheet[above_cell].value
        if above_value and isinstance(above_value, str):
            return f"{above_value.strip()} ({cell_ref})"
    except:
        pass

    # Default label: cell reference and formula snippet
    return f"Formula in {cell_ref}: {formula[:20]}..."

def process_excel_file(input_file, output_file):
    """
    Process an Excel file to extract and convert formulas, then save to CSV.
    Args:
        input_file (str): Path to Excel file
        output_file (str): Path to output CSV file
    """
    # Load workbook
    wb = openpyxl.load_workbook(input_file, data_only=False)
    results = []

    # Iterate through all sheets
    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        for row in sheet.iter_rows():
            for cell in row:
                if cell.data_type == 'f' and cell.value:  # Cell contains a formula
                    formula = cell.value
                    cell_ref = cell.coordinate
                    python_expr = convert_formula_to_python(formula, cell_ref)
                    label = generate_label(cell_ref, sheet, formula)
                    results.append({
                        'Sheet': sheet_name,
                        'Cell': cell_ref,
                        'Original Formula': formula,
                        'Python Expression': python_expr,
                        'Label': label
                    })

    # Save results to CSV
    with open(output_file, 'w', newline='', encoding='utf-8') as f:
        writer = csv.DictWriter(f, fieldnames=['Sheet', 'Cell', 'Original Formula', 'Python Expression', 'Label'])
        writer.writeheader()
        writer.writerows(results)

if __name__ == '__main__':
    input_file = 'input_spreadsheet.xlsx'  # Replace with your Excel file path
    output_file = 'formulas_converted.csv'  # Output CSV file
    process_excel_file(input_file, output_file)
    print(f"Formulas extracted and converted. Results saved to {output_file}")
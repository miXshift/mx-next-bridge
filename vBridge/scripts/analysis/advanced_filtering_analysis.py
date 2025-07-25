#!/usr/bin/env python3
"""
Advanced analysis to find what exact combination would match Excel totals
"""

import pandas as pd
import openpyxl

def main():
    print('🔍 ADVANCED FILTERING ANALYSIS TO MATCH EXCEL TOTALS')
    print('=' * 65)

    # Load Excel campaign list to see exactly which campaigns are included
    wb = openpyxl.load_workbook('data/HydrapakUS_SPABridge_MoM_February 2025vsJanuary 2025.xlsx', data_only=True)
    ws = wb['Campaign']

    excel_campaigns = []
    header_row = 13
    
    # Extract all campaign names from Excel (skip header and total row)
    print('📊 EXTRACTING EXACT CAMPAIGN LIST FROM EXCEL:')
    for row in range(header_row + 2, ws.max_row + 1):  # Start after total row
        campaign = ws.cell(row=row, column=1).value
        jan_spend = ws.cell(row=row, column=2).value
        feb_spend = ws.cell(row=row, column=3).value
        
        if campaign:
            excel_campaigns.append({
                'campaign': str(campaign).strip(),
                'jan_spend': float(jan_spend) if jan_spend else 0,
                'feb_spend': float(feb_spend) if feb_spend else 0
            })

    print(f'  Found {len(excel_campaigns)} campaigns in Excel')
    
    # Calculate totals from Excel campaign list
    excel_jan_calculated = sum(c['jan_spend'] for c in excel_campaigns)
    excel_feb_calculated = sum(c['feb_spend'] for c in excel_campaigns)
    
    print(f'  Excel calculated totals:')
    print(f'    January: ${excel_jan_calculated:,.2f}')
    print(f'    February: ${excel_feb_calculated:,.2f}')

    # Load CSV data
    csv_df = pd.read_csv('output/analyses/mixbridge_jan2025-feb2025_delta_20250721_190317.csv')
    csv_campaigns = csv_df[~csv_df['Campaign'].astype(str).str.contains('Total', case=False, na=False)]
    
    # Create exact filter based on Excel campaign list
    excel_campaign_names = [c['campaign'] for c in excel_campaigns]
    exact_match_filter = csv_campaigns[csv_campaigns['Campaign'].isin(excel_campaign_names)]
    
    csv_jan_exact = exact_match_filter['Spend - January 2025'].sum()
    csv_feb_exact = exact_match_filter['Spend - February 2025'].sum()
    
    print(f'\n🎯 EXACT CAMPAIGN MATCH RESULTS:')
    print(f'  CSV campaigns matching Excel list: {len(exact_match_filter)}')
    print(f'  CSV January total: ${csv_jan_exact:,.2f}')
    print(f'  CSV February total: ${csv_feb_exact:,.2f}')
    
    jan_diff = abs(excel_jan_calculated - csv_jan_exact)
    feb_diff = abs(excel_feb_calculated - csv_feb_exact)
    jan_rel_diff = (jan_diff / excel_jan_calculated * 100) if excel_jan_calculated != 0 else 0
    feb_rel_diff = (feb_diff / excel_feb_calculated * 100) if excel_feb_calculated != 0 else 0
    
    print(f'  Difference from Excel:')
    print(f'    January: ${jan_diff:,.2f} ({jan_rel_diff:.6f}%)')
    print(f'    February: ${feb_diff:,.2f} ({feb_rel_diff:.6f}%)')
    
    # Check if any campaigns in Excel are missing from CSV
    csv_campaign_names = csv_campaigns['Campaign'].tolist()
    missing_from_csv = [name for name in excel_campaign_names if name not in csv_campaign_names]
    
    if missing_from_csv:
        print(f'\n⚠️  CAMPAIGNS IN EXCEL BUT MISSING FROM CSV:')
        for campaign in missing_from_csv:
            print(f'    - {campaign}')
    else:
        print(f'\n✅ All Excel campaigns found in CSV')
    
    # Show sample comparison for individual campaigns
    print(f'\n📋 SAMPLE CAMPAIGN-BY-CAMPAIGN COMPARISON:')
    for i, excel_camp in enumerate(excel_campaigns[:5]):
        campaign_name = excel_camp['campaign']
        excel_jan = excel_camp['jan_spend']
        excel_feb = excel_camp['feb_spend']
        
        csv_match = csv_campaigns[csv_campaigns['Campaign'] == campaign_name]
        if not csv_match.empty:
            csv_jan = csv_match['Spend - January 2025'].iloc[0]
            csv_feb = csv_match['Spend - February 2025'].iloc[0]
            
            jan_match = abs(excel_jan - csv_jan) < 0.01
            feb_match = abs(excel_feb - csv_feb) < 0.01
            
            status = '✅' if jan_match and feb_match else '❌'
            print(f'  {status} {campaign_name[:40]:<40}')
            print(f'      Excel: Jan ${excel_jan:>8.2f} Feb ${excel_feb:>8.2f}')
            print(f'      CSV:   Jan ${csv_jan:>8.2f} Feb ${csv_feb:>8.2f}')
        else:
            print(f'  ❌ {campaign_name[:40]:<40} - NOT FOUND IN CSV')
    
    # Final assessment
    print(f'\n🎯 FINAL ASSESSMENT:')
    tolerance = 0.01
    
    if jan_rel_diff < tolerance and feb_rel_diff < tolerance:
        print(f'  ✅ EXACT MATCH! Campaign name filtering achieves perfect match')
        print(f'     Differences are within {tolerance}% tolerance')
    elif jan_rel_diff < 0.1 and feb_rel_diff < 0.1:
        print(f'  ⚠️  Very close match - within 0.1% tolerance')
        print(f'     Minor discrepancies may be due to rounding or data precision')
    else:
        print(f'  ❌ Significant differences remain even with exact campaign matching')
        print(f'     This suggests different data sources or date ranges')
    
    # Provide the filtering rule
    print(f'\n📝 CAMPAIGN FILTERING RULE FOR EXACT MATCH:')
    print(f'  Include only these {len(excel_campaign_names)} campaigns:')
    for i, name in enumerate(excel_campaign_names[:10]):
        print(f'    {i+1:2}. {name}')
    if len(excel_campaign_names) > 10:
        print(f'    ... and {len(excel_campaign_names) - 10} more campaigns')

if __name__ == "__main__":
    main()